import copy

from agregator.models import UserTasks
from django_celery_results.models import TaskResult
from django.shortcuts import render
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, DEFAULT_FONT, Font


def upload_entity_view(request, tasks_id, entity_type, entity_form, save_func, process_func, error_handler, page):
    user_id = request.user.id
    if request.method == 'POST':
        form = entity_form(request.POST, request.FILES)
        if form.is_valid():
            uploaded_files = form.cleaned_data['files']
            if request.user.is_superuser:
                is_public = True if request.POST['storage_type'] == 'public' else False
            else:
                is_public = False
            entities_ids = save_func(uploaded_files, user_id, is_public)
            task = process_func.apply_async((entities_ids, user_id),
                                            link_error=error_handler.s())
            tasks_id = [task.task_id] + tasks_id
            user_task = UserTasks(user_id=user_id, task_id=task.task_id, files_type=entity_type,
                                  upload_source={'source': 'Пользовательский файл'})
            user_task.save()
            return render(request, page, {'form': form, 'tasks_id': tasks_id})
    else:
        form = entity_form()
    return render(request, page, {'form': form, 'tasks_id': tasks_id})


def get_register_view(request, model, entity_name, public_only_fields=None, private_only_fields=None,
                      template_name=None):
    public = model.objects.filter(is_processing=False, is_public=True).only(*public_only_fields or [])
    private = model.objects.filter(is_processing=False, user_id=request.user.id, is_public=False).only(
        *private_only_fields or [])
    items = public | private
    return render(request, template_name, {entity_name: items})


def generate_excel_report(dataframe, file_path, column_widths=None, height_title=50, height_cell=80):
    with pd.ExcelWriter(file_path) as writer:
        dataframe.to_excel(writer, sheet_name="Sheet1", index=False)

    # Общая настройка стилей Excel
    wb = load_workbook(file_path)
    ws = wb.active

    if column_widths:
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    # Применение стилей
    font = Font(name='Times New Roman', size=11, bold=False, italic=False)
    {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}

    for i in range(1, len(dataframe.values) + 2):
        ws.row_dimensions[i].height = 80 if i == 1 else 100
        for cell in ws[i]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(file_path)
    return file_path


def process_edit_form(request, instance, fields_mapping):
    if request.method == 'POST':
        for field in fields_mapping:
            setattr(instance, field, request.POST[field])
        instance.save()
        return True
    return False


def process_supplement(request, instance):
    input_dict = request.POST.dict()
    supplement = copy.deepcopy(instance.supplement_dict)
    for category, images in instance.supplement_dict.items():
        for i in range(len(images)):
            supplement[category][i]['source'] = input_dict['source-' + images[i]['source']]
            if input_dict['label-' + images[i]['source']]:
                supplement[category][i]['label'] = input_dict['label-' + images[i]['source']]
    return supplement


def create_model_dataframe(model, fields_mapping):
    df = None
    for instance in model.objects.all():
        row = {display: getattr(instance, field) for display, field in fields_mapping.items()}
        df = pd.DataFrame([row]) if df is None else df._append(pd.DataFrame([row]), ignore_index=True)
    return df


def get_scan_task(task_name):
    is_processing = False
    scan_task_id = None

    active_scan_task = TaskResult.objects.filter(
        task_name=task_name
    ).exclude(
        status__in=['SUCCESS', 'FAILURE', 'REVOKED']  # Исключаем точно завершенные
    ).order_by('-date_created').first()
    if active_scan_task:
        scan_task_id = active_scan_task.task_id
        is_processing = True

    return is_processing, scan_task_id, active_scan_task
