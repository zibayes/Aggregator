import simplekml
from django.shortcuts import render, redirect
from .forms import UploadReportsForm, UploadOpenListsForm
from .acts_processing import process_acts, error_handler_acts
from .scientific_reports_processing import process_scientific_reports, error_handler_scientific_reports
from .tech_reports_processing import process_tech_reports, error_handler_tech_reports
from .external_sources import external_sources_processing, external_voan_list_processing
from .open_lists_ocr import process_open_lists, error_handler_open_lists
from .ask import ask_question_with_context
import os
import pandas as pd
import json
import re
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from rest_framework import generics
from urllib.parse import quote
from . import serializers
from django.http import JsonResponse
from django.contrib.auth import login, authenticate
from .forms import UserRegisterForm
from django_celery_results.models import TaskResult
from .models import User, Act, ScientificReport, TechReport, OpenLists, UserTasks, \
    ArchaeologicalHeritageSite, IdentifiedArchaeologicalHeritageSite
from openpyxl import load_workbook
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
from .decorators import owner_or_admin_required
from .files_saving import raw_open_lists_save, raw_reports_save


def get_user_tasks(user_id, file_types, upload_source=False):
    user_tasks = list(UserTasks.objects.filter(user_id=user_id, files_type__in=file_types))
    if upload_source:
        user_tasks = [x for x in user_tasks if x.upload_source['source'] != 'Пользовательский файл']
    else:
        user_tasks = [x for x in user_tasks if x.upload_source['source'] == 'Пользовательский файл']
    user_tasks = [x.task_id for x in user_tasks]
    user_tasks = list(TaskResult.objects.filter(task_id__in=user_tasks).order_by('-date_created'))
    tasks_id = [x.task_id for x in user_tasks]
    return tasks_id


def index(request):
    return render(request, 'index.html')


@login_required
def get_user_tasks_reports(request):
    user = request.user
    tasks_id = get_user_tasks(user.id, ('act', 'scientific_report', 'tech_report'))
    return JsonResponse({'tasks_id': tasks_id})


@login_required
def get_user_tasks_open_lists(request):
    user = request.user
    tasks_id = get_user_tasks(user.id, ('open_list',))
    return JsonResponse({'tasks_id': tasks_id})


@login_required
def get_user_tasks_external(request):
    admin = User.objects.get(is_superuser=True)
    tasks_id = get_user_tasks(admin.id, ('act', 'scientific_report', 'tech_report', 'open_list'), True)
    return JsonResponse({'tasks_id': tasks_id})


@login_required
def deconstructor(request):
    user = request.user
    tasks_id = get_user_tasks(user.id, ('act', 'scientific_report', 'tech_report'))
    if request.method == 'POST':
        if 'acts' in request.POST:
            form = UploadReportsForm()
            return render(request, 'deconstructor.html',
                          {'form': form, 'tasks_id': tasks_id})
        else:
            form = UploadReportsForm(request.POST, request.FILES)
            if form.is_valid():
                uploaded_files = form.cleaned_data['files']
                file_groups = {}
                if 'file_type' in request.POST and 'upload_type' in request.POST:
                    file_type = request.POST['file_type']
                    upload_type = request.POST['upload_type']
                    if user.is_superuser:
                        is_public = True if request.POST['storage_type'] == 'public' else False
                    else:
                        is_public = False
                else:
                    return render(request, 'deconstructor.html', {'form': form, 'tasks_id': tasks_id})
                types_convert = {'текст': 'text', 'приложение': 'images', 'иллюстрации': 'images'}
                if upload_type == 'fully':
                    file_groups['fully'] = []
                    for file in uploaded_files:
                        filename = file.name.lower()
                        report_type = 'all'
                        for typename in types_convert.keys():
                            if typename in filename:
                                index = filename.find(typename)
                                if index >= 0:
                                    report_type = types_convert[typename]
                                    break
                        file_groups['fully'].append({'type': report_type, 'file': file})
                    uploaded_files = []
                elif upload_type == 'mixed':
                    for i in range(len(uploaded_files) - 1, -1, -1):
                        group_name = None
                        filename = uploaded_files[i].name.lower()
                        report_type = 'all'
                        for typename in types_convert.keys():
                            if typename in filename:
                                index = filename.find(typename)
                                if index >= 0:
                                    if index == 0:
                                        group_name = filename[len(typename):]
                                    else:
                                        group_name = filename[:index]
                                    report_type = types_convert[typename]
                                    break
                        if group_name:
                            if group_name in file_groups.keys():
                                file_groups[group_name].append({'type': report_type, 'file': uploaded_files.pop(i)})
                            else:
                                file_groups[group_name] = [{'type': report_type, 'file': uploaded_files.pop(i)}]
                if file_type == 'act':
                    acts_ids = raw_reports_save(file_groups, uploaded_files, Act, user.id, is_public)
                    task = process_acts.apply_async((acts_ids, user.id),
                                                    link_error=error_handler_acts.s())
                elif file_type == 'scientific_report':
                    scientific_reports_ids = raw_reports_save(file_groups, uploaded_files, ScientificReport,
                                                              user.id)
                    task = process_scientific_reports.apply_async((scientific_reports_ids, user.id, is_public),
                                                                  link_error=error_handler_scientific_reports.s())
                elif file_type == 'tech_report':
                    tech_reports_ids = raw_reports_save(file_groups, uploaded_files, TechReport,
                                                        user.id)
                    task = process_tech_reports.apply_async((tech_reports_ids, user.id, is_public),
                                                            link_error=error_handler_tech_reports.s())
                tasks_id = [task.task_id] + tasks_id
                user_task = UserTasks(user_id=user.id, task_id=task.task_id, files_type=file_type,
                                      upload_source={'source': 'Пользовательский файл'})
                user_task.save()

                return render(request, 'deconstructor.html', {'form': form,
                                                              'tasks_id': tasks_id})

    else:
        form = UploadReportsForm()
    return render(request, 'deconstructor.html', {'form': form, 'tasks_id': tasks_id})


@login_required
def external_sources(request):
    is_processing = False
    if request.method == 'POST':
        start_date = end_date = None
        if 'enableDateRange' in request.POST.keys() and 'start_date' in request.POST.keys() and 'start_date' in request.POST.keys():
            start_date = request.POST['start_date']
            end_date = request.POST['end_date']
            match = re.search(r"\d{2}\.\d{2}\.\d{4}", start_date)
            if match:
                date_str = match.group(0)
                start_date = [int(x) for x in date_str.split('.')]
            else:
                start_date = None
            match = re.search(r"\d{2}\.\d{2}\.\d{4}", end_date)
            if match:
                date_str = match.group(0)
                end_date = [int(x) for x in date_str.split('.')]
            else:
                end_date = None
        external_sources_processing.delay(start_date, end_date)
        is_processing = True
    admin = User.objects.get(is_superuser=True)
    tasks_id = get_user_tasks(admin.id, ('act', 'scientific_report', 'tech_report', 'open_list'), True)
    return render(request, 'external_sources.html', {'is_processing': is_processing, 'tasks_id': tasks_id})


def constructor(request):
    return render(request, 'constructor.html')


def interactive_map(request):
    acts = Act.objects.filter(is_processing=False)
    scientific_reports = ScientificReport.objects.filter(is_processing=False)
    tech_report = TechReport.objects.filter(is_processing=False)
    all_coordinates = {'Акты': {}, 'Научные отчёты': {}, 'Научно-технические отчёты': {}}
    for act in acts:
        all_coordinates['Акты'][
            act.source[0]['origin_filename']] = act.coordinates  # TODO: Подобрать более удачный нейминг?
    for report in scientific_reports:
        all_coordinates['Научные отчёты'][report.source[0]['origin_filename']] = report.coordinates
    for report in tech_report:
        all_coordinates['Научно-технические отчёты'][report.source[0]['origin_filename']] = report.coordinates
    return render(request, 'interactive_map.html', {'all_coordinates': all_coordinates})


def download_all_coordinates(request):
    if request.method == 'POST':
        acts = Act.objects.filter(is_processing=False)
        scientific_reports = ScientificReport.objects.filter(is_processing=False)
        tech_report = TechReport.objects.filter(is_processing=False)
        all_coordinates = {'Акты': {}, 'Научные отчёты': {}, 'Научно-технические отчёты': {}}
        coordinates_to_download = {}

        for act in acts:
            all_coordinates['Акты'][
                act.source[0]['origin_filename']] = act.coordinates  # TODO: Подобрать более удачный нейминг?
        for report in scientific_reports:
            all_coordinates['Научные отчёты'][report.source[0]['origin_filename']] = report.coordinates
        for report in tech_report:
            all_coordinates['Научно-технические отчёты'][report.source[0]['origin_filename']] = report.coordinates

        for report_type, reports in all_coordinates.items():
            for report, groups in reports.items():
                for group, point in groups.items():
                    for point_name, coords in point.items():
                        for key in request.POST.keys():
                            if f'{report_type}-{report}-{group}-{point_name}' == key:
                                if report_type not in coordinates_to_download.keys():
                                    coordinates_to_download[report_type] = {}
                                if report not in coordinates_to_download[report_type].keys():
                                    coordinates_to_download[report_type][report] = {}
                                if group not in coordinates_to_download[report_type][report].keys():
                                    coordinates_to_download[report_type][report][group] = {}
                                coordinates_to_download[report_type][report][group][point_name] = coords

        if coordinates_to_download:
            kml = simplekml.Kml()
            catalog_style = simplekml.Style()
            catalog_style.iconstyle.color = simplekml.Color.blue
            photos_style = simplekml.Style()
            photos_style.iconstyle.color = simplekml.Color.green
            pits_style = simplekml.Style()
            pits_style.iconstyle.color = simplekml.Color.red
            current_style = current_group = None
            for report_type, reports in coordinates_to_download.items():
                report_type_folder = kml.newfolder(name=report_type)
                for report, groups in reports.items():
                    report_folder = report_type_folder.newfolder(name=report)
                    for group, point in groups.items():
                        system_check = True  # 'WGS-84' in group or 'WGS84' in group or 'WGS 84' in group or 'Шурф' in group
                        if 'фотофиксации' in group:
                            current_style = photos_style
                            photos_group = report_folder.newfolder(name=group)
                            current_group = photos_group
                        elif 'Каталог' in group:
                            current_style = catalog_style
                            catalog_group = report_folder.newfolder(name=group)
                            current_group = catalog_group
                        elif 'Шурфы' in group:
                            current_style = pits_style
                            pits_group = report_folder.newfolder(name=group)
                            current_group = pits_group
                        for point_name, coords in point.items():
                            if current_group and system_check:
                                photo_point = current_group.newpoint(name=str(point_name),
                                                                     coords=[
                                                                         (coords[1],
                                                                          coords[
                                                                              0])])  # TODO: менять их местами или нет?!
                                photo_point.style = current_style

            file_path = f'uploaded_files/{request.user.id}-all_coordinates.kml'
            kml.save(file_path)
            return redirect('/' + file_path)
        return JsonResponse({'response': f'There is no selected coordinates'})
    return JsonResponse({'response': f'Method {request.method} is not available for this URL'})


def acts_register(request):
    acts_public = Act.objects.filter(is_processing=False, is_public=True).only('id', 'user_id', 'date_uploaded',
                                                                               'is_processing', 'year', 'finish_date',
                                                                               'type', 'name_number', 'place',
                                                                               'customer', 'area',
                                                                               'expert', 'executioner', 'open_list',
                                                                               'conclusion',
                                                                               'border_objects', 'source')
    acts_private = Act.objects.filter(is_processing=False,
                                      user_id=request.user.id, is_public=False).only('id', 'user_id', 'date_uploaded',
                                                                                     'is_processing', 'year',
                                                                                     'finish_date',
                                                                                     'type', 'name_number', 'place',
                                                                                     'customer', 'area',
                                                                                     'expert', 'executioner',
                                                                                     'open_list', 'conclusion',
                                                                                     'border_objects', 'source')
    acts = acts_public | acts_private
    return render(request, 'acts_register.html', {'acts': acts})


@login_required
def open_list_ocr(request):
    user_id = request.user.id
    tasks_id = get_user_tasks(user_id, ('open_list',))
    if request.method == 'POST':
        form = UploadOpenListsForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_files = form.cleaned_data['files']
            if request.user.is_superuser:
                is_public = True if request.POST['storage_type'] == 'public' else False
            else:
                is_public = False
            open_lists_ids = raw_open_lists_save(uploaded_files, user_id, is_public)
            task = process_open_lists.apply_async((open_lists_ids, user_id),
                                                  link_error=error_handler_open_lists.s())
            tasks_id = [task.task_id] + tasks_id
            user_task = UserTasks(user_id=user_id, task_id=task.task_id, files_type='open_list',
                                  upload_source={'source': 'Пользовательский файл'})
            user_task.save()
            return render(request, 'open_list_ocr.html', {'form': form, 'tasks_id': tasks_id})
    else:
        form = UploadOpenListsForm()
    return render(request, 'open_list_ocr.html', {'form': form, 'tasks_id': tasks_id})


def gpt_chat(request):
    return render(request, 'gpt_chat.html')


def ask_gpt(request):
    if request.method == 'POST':
        body_unicode = request.body.decode('utf-8')
        body_data = json.loads(body_unicode)
        answer = ask_question_with_context(body_data['messages'][1]['content'])
        return JsonResponse({'choices': [{'message': {'content': answer}}]})


def user_register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Автоматически авторизуем пользователя после регистрации
            return redirect('profile')
        else:
            return render(request, 'register.html', {'error': 'Неверные учетные данные'})
    else:
        form = UserRegisterForm()
    return render(request, 'register.html', {'form': form})


def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('profile')
        else:
            return render(request, 'login.html', {'error': 'Неверные учетные данные'})
    return render(request, 'login.html')


@login_required
def custom_logout(request):
    logout(request)
    return redirect('index')


@login_required
def profile(request):
    return render(request, 'profile.html', {'user_to_show': request.user})


@login_required
def settings(request):
    if request.method == 'POST':
        user = request.user
        user.avatar = request.FILES['avatar']
        user.first_name = request.POST['first_name']
        user.last_name = request.POST['last_name']
        user.username = request.POST['username']
        user.email = request.POST['email']
        password = request.POST['password']

        if password:
            user.set_password(password)  # Хешируйте новый пароль
            update_session_auth_hash(request, user)  # Убедитесь, что сеанс остается активным
        user.save()
        messages.success(request, 'Профиль успешно обновлен.')
        return redirect('profile')  # Перенаправление на страницу профиля

    return render(request, 'settings.html')


def open_lists_register(request):
    open_lists_public = OpenLists.objects.filter(is_processing=False, is_public=True)
    open_lists_private = OpenLists.objects.filter(is_processing=False, user_id=request.user.id, is_public=False)
    open_lists = open_lists_public | open_lists_private
    return render(request, 'open_lists_register.html', {'open_lists': open_lists})


def open_lists_register_download(request):
    table_path = "uploaded_files/open_lists/Открытые листы.xlsx"
    open_lists = OpenLists.objects.all()
    if not open_lists:
        return redirect(open_lists_register)
    df_existing = None
    for list in open_lists:
        list_data = {'Номер листа': '', 'Держатель': '', 'Объект': '', 'Работы': '', 'Начало срока': '',
                     'Конец срока': ''}
        list_data['Номер листа'] = list.number,
        list_data['Держатель'] = list.holder,
        list_data['Объект'] = list.object,
        list_data['Работы'] = list.works,
        list_data['Начало срока'] = list.start_date,
        list_data['Конец срока'] = list.end_date,
        df_new = pd.DataFrame(list_data, columns=list_data.keys(), index=[0])
        if df_existing is None:
            df_existing = df_new
        else:
            df_existing = df_existing._append(df_new, ignore_index=True)
    with pd.ExcelWriter(table_path) as writer:
        df_existing.to_excel(writer, sheet_name="Sheet1", index=False)
    wb = load_workbook(table_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 100
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}
    for i in range(1, len(df_existing.values) + 2):
        if i == 1:
            ws.row_dimensions[0].height = 50
        else:
            ws.row_dimensions[i].height = 80
        for cell in ws[i]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(table_path)
    return redirect('/uploaded_files/open_lists/Открытые листы.xlsx')


def acts_register_download(request):
    table_path = "uploaded_files/acts/РЕЕСТР актов ГИКЭ.xlsx"
    act_parts = ['ГОД', 'Дата окончания проведения ГИКЭ', 'Вид ГИКЭ', 'Номер (если имеется) и наименование Акта ГИКЭ',
                 'Место проведения экспертизы',
                 'Заказчик работ (*если не указан, то заказчик экспертизы)',
                 'Площадь, протяжённость и/или др. параменты объекта', 'Эксперт (физ. или юр.лицо)',
                 'Исполнитель полевых работ (юр. лицо)', 'ОЛ', 'Заключение. Выявленые объекты.',
                 'Объекты расположенные в непосредственной близости. Для границ']
    acts = Act.objects.all()
    if not acts:
        return redirect(open_lists_register)
    df_existing = None
    for act in acts:
        act_parts_info = {i: '' for i in act_parts}
        act_parts_info['ГОД'] = act.year,
        act_parts_info['Дата окончания проведения ГИКЭ'] = act.finish_date,
        act_parts_info['Вид ГИКЭ'] = act.type,
        act_parts_info['Номер (если имеется) и наименование Акта ГИКЭ'] = act.name_number,
        act_parts_info['Место проведения экспертизы'] = act.place,
        act_parts_info['Заказчик работ (*если не указан, то заказчик экспертизы)'] = act.customer,
        act_parts_info['Площадь, протяжённость и/или др. параменты объекта'] = act.area,
        act_parts_info['Эксперт (физ. или юр.лицо)'] = act.expert,
        act_parts_info['Исполнитель полевых работ (юр. лицо)'] = act.executioner,
        act_parts_info['ОЛ'] = act.open_list,
        act_parts_info['Заключение. Выявленые объекты.'] = act.conclusion,
        act_parts_info['Объекты расположенные в непосредственной близости. Для границ'] = act.border_objects,
        df_new = pd.DataFrame(act_parts_info, columns=act_parts_info.keys(), index=[0])
        if df_existing is None:
            df_existing = df_new
        else:
            df_existing = df_existing._append(df_new, ignore_index=True)
    with pd.ExcelWriter(table_path) as writer:
        df_existing.to_excel(writer, sheet_name="Sheet1", index=False)
    wb = load_workbook(table_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 6.86
    ws.column_dimensions['B'].width = 10.14
    ws.column_dimensions['C'].width = 10.14
    ws.column_dimensions['D'].width = 66.43
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 26
    ws.column_dimensions['G'].width = 20.71
    ws.column_dimensions['H'].width = 18.43
    ws.column_dimensions['I'].width = 24.71
    ws.column_dimensions['J'].width = 21.29
    ws.column_dimensions['K'].width = 26
    ws.column_dimensions['L'].width = 27.29
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}
    for i in range(1, len(df_existing.values) + 2):
        if i == 1:
            ws.row_dimensions[0].height = 50
        else:
            ws.row_dimensions[i].height = 80
        for cell in ws[i]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(table_path)
    return redirect('/uploaded_files/acts/РЕЕСТР актов ГИКЭ.xlsx')


def scientific_reports_register_download(request):
    table_path = "uploaded_files/scientific_reports/РЕЕСТР ПНО.xlsx"
    table_columns = ['Год написания отчёта', 'Название отчёта', 'Организация', 'Автор',
                     'Открытый лист', 'Населённый пункт',
                     'Вид работ', 'Площадь', 'Исполнители',
                     'Заключение']
    reports = ScientificReport.objects.all()
    if not reports:
        return redirect(scientific_reports_register)
    df_existing = None
    for report in reports:
        table_columns_info = {i: '' for i in table_columns}
        table_columns_info['Год написания отчёта'] = report.writing_date
        table_columns_info['Название отчёта'] = report.name
        table_columns_info['Организация'] = report.organization
        table_columns_info['Автор'] = report.author
        table_columns_info['Открытый лист'] = report.open_list
        table_columns_info['Населённый пункт'] = report.place
        table_columns_info['Исполнители'] = report.contractors
        table_columns_info['Площадь'] = report.area_info
        df_new = pd.DataFrame(table_columns_info, columns=table_columns_info.keys(), index=[0])
        if df_existing is None:
            df_existing = df_new
        else:
            df_existing = df_existing._append(df_new, ignore_index=True)
    with pd.ExcelWriter(table_path) as writer:
        df_existing.to_excel(writer, sheet_name="Sheet1", index=False)
    wb = load_workbook(table_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 26
    ws.column_dimensions['G'].width = 20.71
    ws.column_dimensions['H'].width = 18.43
    ws.column_dimensions['I'].width = 24.71
    ws.column_dimensions['J'].width = 21.29
    ws.column_dimensions['K'].width = 26
    ws.column_dimensions['L'].width = 27.29
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}
    for i in range(1, len(df_existing.values) + 2):
        if i == 1:
            ws.row_dimensions[0].height = 50
        else:
            ws.row_dimensions[i].height = 80
        for cell in ws[i]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(table_path)
    return redirect('/uploaded_files/scientific_reports/РЕЕСТР ПНО.xlsx')


def tech_reports_register_download(request):
    table_path = "uploaded_files/tech_reports/РЕЕСТР ПНТО.xlsx"
    table_columns = ['Год написания отчёта', 'Название отчёта', 'Организация', 'Автор',
                     'Открытый лист', 'Населённый пункт',
                     'Вид работ', 'Площадь', 'Исполнители',
                     'Заключение']
    reports = TechReport.objects.all()
    if not reports:
        return redirect(scientific_reports_register)
    df_existing = None
    for report in reports:
        table_columns_info = {i: '' for i in table_columns}
        table_columns_info['Год написания отчёта'] = report.writing_date
        table_columns_info['Название отчёта'] = report.name
        table_columns_info['Организация'] = report.organization
        table_columns_info['Автор'] = report.author
        table_columns_info['Открытый лист'] = report.open_list
        table_columns_info['Населённый пункт'] = report.place
        table_columns_info['Исполнители'] = report.contractors
        table_columns_info['Площадь'] = report.area_info
        df_new = pd.DataFrame(table_columns_info, columns=table_columns_info.keys(), index=[0])
        if df_existing is None:
            df_existing = df_new
        else:
            df_existing = df_existing._append(df_new, ignore_index=True)
    with pd.ExcelWriter(table_path) as writer:
        df_existing.to_excel(writer, sheet_name="Sheet1", index=False)
    wb = load_workbook(table_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 26
    ws.column_dimensions['G'].width = 20.71
    ws.column_dimensions['H'].width = 18.43
    ws.column_dimensions['I'].width = 24.71
    ws.column_dimensions['J'].width = 21.29
    ws.column_dimensions['K'].width = 26
    ws.column_dimensions['L'].width = 27.29
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}
    for i in range(1, len(df_existing.values) + 2):
        if i == 1:
            ws.row_dimensions[0].height = 50
        else:
            ws.row_dimensions[i].height = 80
        for cell in ws[i]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(table_path)
    return redirect("/uploaded_files/tech_reports/РЕЕСТР ПНТО.xlsx")


def scientific_reports_register(request):
    reports_public = ScientificReport.objects.filter(is_processing=False, is_public=True).only('id', 'user_id',
                                                                                               'date_uploaded',
                                                                                               'upload_source',
                                                                                               'is_processing', 'name',
                                                                                               'organization', 'author',
                                                                                               'open_list',
                                                                                               'writing_date',
                                                                                               'contractors', 'source',
                                                                                               'place', 'area_info',
                                                                                               'results', 'conclusion')
    reports_private = ScientificReport.objects.filter(is_processing=False, user_id=request.user.id,
                                                      is_public=False).only('id', 'user_id', 'date_uploaded',
                                                                            'upload_source', 'is_processing', 'name',
                                                                            'organization', 'author', 'open_list',
                                                                            'writing_date', 'contractors', 'source',
                                                                            'place', 'area_info', 'results',
                                                                            'conclusion')
    reports = reports_public | reports_private
    return render(request, 'scientific_reports_register.html', {'reports': reports})


def tech_reports_register(request):
    reports_public = TechReport.objects.filter(is_processing=False, is_public=True).only('id', 'user_id',
                                                                                         'date_uploaded',
                                                                                         'upload_source',
                                                                                         'is_processing', 'name',
                                                                                         'organization', 'author',
                                                                                         'open_list',
                                                                                         'writing_date', 'contractors',
                                                                                         'source',
                                                                                         'place', 'area_info',
                                                                                         'results', 'conclusion')
    reports_private = TechReport.objects.filter(is_processing=False, user_id=request.user.id, is_public=False).only(
        'id', 'user_id', 'date_uploaded',
        'upload_source', 'is_processing', 'name',
        'organization', 'author', 'open_list',
        'writing_date', 'contractors', 'source',
        'place', 'area_info', 'results', 'conclusion')
    reports = reports_public | reports_private
    return render(request, 'tech_reports_register.html', {'reports': reports})


def users(request, pk):
    user = User.objects.get(id=pk)
    return render(request, 'profile.html', {'user_to_show': user})


def map(request, report_type, pk):
    report = None
    if report_type == 'act':
        report = Act.objects.get(id=pk)
    elif report_type == 'scientific_report':
        report = ScientificReport.objects.get(id=pk)
    elif report_type == 'tech_report':
        report = TechReport.objects.get(id=pk)
    report_name = report.source[0]['origin_filename']
    coordinates = report.coordinates if report else {}
    return render(request, 'interactive_map.html',
                  {'coordinates': coordinates, 'report_type': report_type, 'pk': pk, 'report_name': report_name})


def download_coordinates(request, report_type, pk):
    if request.method == 'POST':
        report = None
        if report_type == 'act':
            report = Act.objects.get(id=pk)
        elif report_type == 'scientific_report':
            report = ScientificReport.objects.get(id=pk)
        elif report_type == 'tech_report':
            report = TechReport.objects.get(id=pk)
        coordinates = report.coordinates if report else {}
        coordinates_to_download = {}

        for group, point in coordinates.items():
            for point_name, coords in point.items():
                if f'{group}-{point_name}' in request.POST.keys():
                    if group not in coordinates_to_download.keys():
                        coordinates_to_download[group] = {}
                    coordinates_to_download[group][point_name] = coords

        if coordinates_to_download:
            kml = simplekml.Kml()
            catalog_style = simplekml.Style()
            catalog_style.iconstyle.color = simplekml.Color.blue
            photos_style = simplekml.Style()
            photos_style.iconstyle.color = simplekml.Color.green
            pits_style = simplekml.Style()
            pits_style.iconstyle.color = simplekml.Color.red
            current_style = current_group = None
            for group, point in coordinates_to_download.items():
                system_check = True  # 'WGS-84' in group or 'WGS84' in group or 'WGS 84' in group or 'Шурф' in group
                if 'фотофиксации' in group:
                    current_style = photos_style
                    photos_group = kml.newfolder(name=group)
                    current_group = photos_group
                elif 'Каталог' in group:
                    current_style = catalog_style
                    catalog_group = kml.newfolder(name=group)
                    current_group = catalog_group
                elif 'Шурфы' in group:
                    current_style = pits_style
                    pits_group = kml.newfolder(name=group)
                    current_group = pits_group
                for point_name, coords in point.items():
                    if current_group and system_check:
                        photo_point = current_group.newpoint(name=str(point_name),
                                                             coords=[
                                                                 (coords[1],
                                                                  coords[0])])  # TODO: менять их местами или нет?!
                        photo_point.style = current_style

            file_path = f'uploaded_files/{report_type}s/{pk}_{report_type}/{pk}_{report_type}/coordinates.kml'
            kml.save(file_path)
            return redirect('/' + file_path)
        return JsonResponse({'response': f'There is no selected coordinates'})
    return JsonResponse({'response': f'Method {request.method} is not available for this URL'})


def acts(request, pk):
    act = Act.objects.get(id=pk)
    return render(request, 'act.html', {'act': act})


@login_required
@owner_or_admin_required(Act)
def acts_edit(request, pk):
    act = Act.objects.get(id=pk)
    if request.method == 'POST':
        act.year = request.POST['year']
        act.finish_date = request.POST['finish_date']
        act.type = request.POST['type']
        act.name_number = request.POST['name_number']
        act.place = request.POST['place']
        act.customer = request.POST['customer']
        act.area = request.POST['area']
        act.expert = request.POST['expert']
        act.executioner = request.POST['executioner']
        act.open_list = request.POST['open_list']
        act.conclusion = request.POST['conclusion']
        act.border_objects = request.POST['border_objects']
        act.save()
        messages.success(request, 'Акт успешно обновлен.')
        return redirect(f'/acts/{act.id}')  # Перенаправление на страницу профиля

    return render(request, 'act_edit.html', {'act': act})


@login_required
@owner_or_admin_required(Act)
def acts_delete(request, pk):
    act_instance = Act.objects.get(id=pk)
    act_instance.delete()
    return redirect(f'acts_register')


@login_required
# @owner_or_admin_required(UserTasks)
def download_delete(request, task_id):
    user_task = UserTasks.objects.get(task_id=task_id)
    user_task.delete()
    task = TaskResult.objects.get(task_id=task_id)
    task.delete()
    return JsonResponse({'response': 'deleted'})


def scientific_reports(request, pk):
    report = ScientificReport.objects.get(id=pk)
    return render(request, 'scientific_report.html', {'report': report})


@login_required
@owner_or_admin_required(ScientificReport)
def scientific_reports_edit(request, pk):
    report = ScientificReport.objects.get(id=pk)
    if request.method == 'POST':
        report.name = request.POST['name']
        report.organization = request.POST['organization']
        report.author = request.POST['author']
        report.open_list = request.POST['open_list']
        report.writing_date = request.POST['writing_date']
        report.introduction = request.POST['introduction']
        report.contractors = request.POST['contractors']
        report.place = request.POST['place']
        report.area_info = request.POST['area_info']
        report.research_history = request.POST['research_history']
        report.results = request.POST['results']
        report.conclusion = request.POST['conclusion']
        report.save()
        messages.success(request, 'Отчёт успешно обновлен.')
        return redirect(f'/scientific_reports/{report.id}')

    return render(request, 'scientific_report_edit.html', {'report': report})


@login_required
@owner_or_admin_required(ScientificReport)
def scientific_reports_delete(request, pk):
    report_instance = ScientificReport.objects.get(id=pk)
    report_instance.delete()
    return redirect(f'scientific_reports_register')


def tech_reports(request, pk):
    report = TechReport.objects.get(id=pk)
    return render(request, 'tech_report.html', {'report': report})


@login_required
@owner_or_admin_required(TechReport)
def tech_reports_edit(request, pk):
    report = TechReport.objects.get(id=pk)
    if request.method == 'POST':
        report.name = request.POST['name']
        report.organization = request.POST['organization']
        report.author = request.POST['author']
        report.open_list = request.POST['open_list']
        report.writing_date = request.POST['writing_date']
        report.introduction = request.POST['introduction']
        report.contractors = request.POST['contractors']
        report.place = request.POST['place']
        report.area_info = request.POST['area_info']
        report.research_history = request.POST['research_history']
        report.results = request.POST['results']
        report.conclusion = request.POST['conclusion']
        report.save()
        messages.success(request, 'Отчёт успешно обновлен.')
        return redirect(f'/tech_reports/{report.id}')

    return render(request, 'tech_report_edit.html', {'report': report})


@login_required
@owner_or_admin_required(TechReport)
def tech_reports_delete(request, pk):
    report_instance = TechReport.objects.get(id=pk)
    report_instance.delete()
    return redirect(f'tech_reports_register')


def open_lists(request, pk):
    open_list = OpenLists.objects.get(id=pk)
    return render(request, 'open_list.html', {'open_list': open_list})


@login_required
@owner_or_admin_required(OpenLists)
def open_lists_edit(request, pk):
    open_list = OpenLists.objects.get(id=pk)
    if request.method == 'POST':
        open_list.number = request.POST['number']
        open_list.holder = request.POST['holder']
        open_list.object = request.POST['object']
        open_list.works = request.POST['works']
        open_list.start_date = request.POST['start_date']
        open_list.end_date = request.POST['end_date']
        open_list.save()
        messages.success(request, 'Открытый лист успешно обновлен.')
        return redirect(f'/open_lists/{open_list.id}')

    return render(request, 'open_list_edit.html', {'open_list': open_list})


@login_required
@owner_or_admin_required(OpenLists)
def open_lists_delete(request, pk):
    list_instance = OpenLists.objects.get(id=pk)
    list_instance.delete()
    return redirect(f'open_lists_register')


def archaeological_heritage_sites(request):
    oan = ArchaeologicalHeritageSite.objects.all()
    return render(request, 'archaeological_heritage_site.html', {'oan': oan})


def identified_archaeological_heritage_sites(request):
    voan = IdentifiedArchaeologicalHeritageSite.objects.all()
    return render(request, 'identified_archaeological_heritage_site.html', {'voan': voan})


def archaeological_heritage_sites_download(request):
    current_lists = 'uploaded_files/voan_list/current_lists.txt'
    link = None
    with open(current_lists, 'r') as file:
        for line in file.readlines():
            if 'list_voan - ' in line:
                link = line.replace('list_voan - ', '').strip()
    if link is None:
        return redirect(request.META.get('HTTP_REFERER', '/'))
    return redirect('/' + quote(link))


def identified_archaeological_heritage_sites_download(request):
    current_lists = 'uploaded_files/voan_list/current_lists.txt'
    link = None
    with open(current_lists, 'r') as file:
        for line in file.readlines():
            if 'list_oan - ' in line:
                link = line.replace('list_oan - ', '').strip()
    if link is None:
        return redirect(request.META.get('HTTP_REFERER', '/'))
    return redirect('/' + quote(link))


def update_voan_list(request):
    external_voan_list_processing()
    return redirect(request.META.get('HTTP_REFERER', '/'))


class UserList(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = serializers.UserSerializer


class UserDetail(generics.RetrieveAPIView):
    queryset = User.objects.all()
    serializer_class = serializers.UserSerializer


class ActList(generics.ListAPIView):
    queryset = Act.objects.all()
    serializer_class = serializers.ActSerializer


class ActDetail(generics.RetrieveAPIView):
    queryset = Act.objects.all()
    serializer_class = serializers.ActSerializer


class ScientificReportList(generics.ListAPIView):
    queryset = ScientificReport.objects.all()
    serializer_class = serializers.ScientificReport


class ScientificReportDetail(generics.RetrieveAPIView):
    queryset = ScientificReport.objects.all()
    serializer_class = serializers.ScientificReportSerializer


class TechReportList(generics.ListAPIView):
    queryset = TechReport.objects.all()
    serializer_class = serializers.TechReport


class TechReportDetail(generics.RetrieveAPIView):
    queryset = TechReport.objects.all()
    serializer_class = serializers.TechReportSerializer
