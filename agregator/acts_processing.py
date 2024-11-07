import fitz  # PyMuPDF
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
import os
from UliPlot.XLSX import auto_adjust_xlsx_column_width
import time
from datetime import datetime, date
import math
import numpy as np


SQUARE_RESERVE = []


def choose_pdf_file() -> str:
    # Открываем окно выбора файла
    # file_path = filedialog.askopenfilename(title="Выберите PDF файл", filetypes=[("PDF файлы", "*.pdf")])
    file_path = filedialog.askdirectory(title="Выберите папку")
    if file_path:
        return file_path


def get_gike_object_size(text_to_write: str, table_info: dict) -> None:
    attr_filled = 'Площадь, протяжённость и/или др. параменты объекта' in table_info.keys()
    if not attr_filled or attr_filled and 'Общ. S' not in table_info[
        'Площадь, протяжённость и/или др. параменты объекта']:
        square = re.search(r'Общ.+[ \n]+площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*', text_to_write, re.IGNORECASE)
        if not square:
            square = re.search(r'площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*', text_to_write,
                               re.IGNORECASE)
        # общ.+[ \n]+площадь.*\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+
        if square:
            square = re.search(r'\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*', square.group(0), re.IGNORECASE).group(0)
            if 'га ' in square:
                square = square.strip()[:square.rfind('га ')+2]
            if 'кв. м' in square or not re.search(r'[А-Яа-я.]+', square, re.IGNORECASE):
                SQUARE_RESERVE.append(square)
            else:
                table_info['Площадь, протяжённость и/или др. параменты объекта'] = 'Общ. S = ' + square
    if not attr_filled or attr_filled and 'протяж.' not in table_info[
        'Площадь, протяжённость и/или др. параменты объекта']:
        length = re.search(r'протяж.*\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+', text_to_write, re.IGNORECASE)
        if length:
            length = re.search(r'\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+', length.group(0), re.IGNORECASE).group(0)
            if 'Площадь, протяжённость и/или др. параменты объекта' not in table_info.keys():
                table_info['Площадь, протяжённость и/или др. параменты объекта'] = 'протяж. ' + length
            else:
                table_info[
                    'Площадь, протяжённость и/или др. параменты объекта'] += '\nпротяж. ' + length
    if not attr_filled or attr_filled and 'S лин.' not in table_info[
        'Площадь, протяжённость и/или др. параменты объекта']:
        square_line = re.search(r'площ[а-яА-ЯёЁ]+[ \n]+лин.*\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+', text_to_write, re.IGNORECASE)
        if square_line:
            square_line = re.search(r'\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+', square_line.group(0), re.IGNORECASE).group(0)
            table_info['Площадь, протяжённость и/или др. параменты объекта'] += ' (S лин. ЗУ = ' + square_line + ')'


def extract_text_and_images(pdf_file):
    # Открываем PDF-файл
    document = fitz.open(pdf_file)
    folder = pdf_file[:pdf_file.rfind(".")]
    Path(folder).mkdir(exist_ok=True)

    # Разделы
    act_parts = ['Акт', 'Дата начала', 'Дата окончания', # проведения экспертизы
                    r'\d*\.*.*Место проведения [экспертизы]*:*', r'\d*\.*.*Заказчик экспертизы', r'\d*\.*.*[Сведения об]* эксперте',
                    'Отношени[яе]+ к заказчику', 'Цель экспертизы:', 'Объект .*?[экспертизы]*:*',  # Объект экспертизы:*
                    'Перечень документов, представленных', 'Сведения о проведенных исследованиях',
                    'Факты и сведения, выявленные .*\n*.*исследований', # 'Факты и сведения, выявленные и установленные в результате проведенных исследований', 'Координаты',
                    'Перечень[а-яА-ЯёЁ \n,]*литературы',
                    # 'Перечень документов и материалов, собранных и полученных при проведении '
                    #              'экспертизы, а также использованной для нее специальной, технической и '
                    #              'справочной литературы', 'Обоснования вывода экспертизы',
                    'Вывод экспертизы', 'Перечень приложений']
    act_sub_parts = ['Характеристика объекта']

    table_info = {}
    table_columns = ['ГОД',	'Дата окончания проведения ГИКЭ', 'Вид ГИКЭ', 'Номер (если имеется) и наименование Акта ГИКЭ',
                     'Место проведения экспертизы', # 'Муниципальный район или муниципальный округ'
                     'Заказчик работ (*если не указан, то заказчик экспертизы)',
                     'Площадь, протяжённость и/или др. параменты объекта', 'Эксперт (физ. или юр.лицо)',
                     'Исполнитель полевых работ (юр. лицо)', 'ОЛ', 'Заключение. Выявленые объекты.',
                     'Объекты расположенные в непосредственной близости. Для границ']
    months = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06', 'июля': '07',
              'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12',}
    broken_structure = False
    exploration_object= False
    sectors_square = []
    text_reserve = None
    voan_reserve = None
    several_experts = False
    full_name = False

    # Создаем или очищаем текстовый файл
    with open(folder + "/" + "text.txt", "w", encoding="utf-8") as text_file:
        extracted_images = []
        current_part = 0
        for page_number in range(len(document)):
            page = document[page_number]
            # Извлечение текста
            text = page.get_text()

            if act_parts[current_part] == 'Акт':
                act = re.search(r'А *К *Т *№* *\d*/*\d*(?!.*подписан).*', text, re.IGNORECASE)
                # А *К *Т *№* *\d*/*\d*\n*(?!.*подписан).*\n*.*
                # А *К *Т № \d+/*\d*\n*.*
                if act:
                    text_to_write = act.group(0)
                else:
                    text_to_write = text[re.search(act_parts[current_part], text, re.IGNORECASE).start():re.search('Настоящий Акт', text, re.IGNORECASE).start()]
                text_file.write(f"--- {act_parts[current_part]} --- (стр. {page_number + 1}):\n{text_to_write}\n")
                current_part += 1
                if '№' not in text_to_write:
                    text_to_write += " б/н"
                table_info['Номер (если имеется) и наименование Акта ГИКЭ'] = text_to_write

            while True:
                current_index = re.search(act_parts[current_part].replace(' ', r'[ \n]*'), text)
                if current_index:
                    current_index = current_index.end()
                else:
                    current_index = 0

                next_index = None
                if current_part + 1 < len(act_parts):
                    next_index = re.search(act_parts[current_part+1].replace(' ', r'[ \n]*'), text)

                text_to_write = text[current_index:next_index.start() if next_index else len(text)]
                text_file.write(f"--- {act_parts[current_part]} --- (стр. {page_number + 1}):\n{text_to_write}\n")

                full_time_interval = None
                interval_type = None
                if act_parts[current_part] == 'Дата начала':
                    full_time_interval = re.search(r'период с \d{2}.\d{2}.\d{4}[ \n]+[г. \n]*по[ \n]+\d{2}.\d{2}.\d{4}[ \n]+г.', text_to_write)
                    # период с \d+.\d+.\d+[ \n]+г.[ \n]+по[ \n]+\d+.\d+.\d+[ \n]+г.
                    if not full_time_interval:
                        full_time_interval = re.search(r'период с «*\d+»* [А-Яа-яёЁ]+ \d+ г\.*.*[ \n]+по[ \n]+«*\d+»* [А-Яа-яёЁ]+ \d+ г\.*',
                                         text_to_write)
                        interval_type = 'words'
                    else:
                        interval_type = 'dots'

                if act_parts[current_part] == 'Дата окончания' or full_time_interval:
                    text_to_write = text[re.search(r'Дата начала', text, re.IGNORECASE).end():]
                    if full_time_interval and interval_type == 'dots':
                        date = full_time_interval
                        current_part += 2
                        date = re.search(r'по[ \n]+\d{2}.\d{2}.\d{4}[ \n]+г', date.group(0)).group(0)
                    else:
                        date = re.findall(r'\d+ *\d+ *\.\d{2} *\d*\.\d{4} *\d* *г*', text_to_write) # (\d )*\d+ *\.\d+ *\d*\.\d+ *\d* *г*  //  \d+\.\d+\.\d+ *г*\.*  //  \d+ *\.\d+ *\d*\.\d+ *\d* *г*
                        if date:
                            if len(date) > 1:
                                date = date[1]
                            else:
                                date = date[0]
                        date_words = re.findall(r'«*\d+»* *[А-Яа-яёЁ]+ \d+ г\.*', text_to_write) # TODO: rework?
                        if date_words:
                            if len(date_words) > 1:
                                date_words = date_words[1]
                            else:
                                date_words = date_words[0]
                        if date and date_words and (text_to_write.find(date) > text_to_write.find(date_words) or 'Постнов' in text_to_write): # TODO: people style?
                            date = None
                    if date and interval_type != 'words':
                        date = date.replace('по ', '').replace(' ', '')
                        year = date[date.rfind('.')+1:]
                        if 'г' in year:
                            year = year[:year.rfind('г')]
                        table_info['ГОД'] = year
                        index = date.rfind(' ')
                        table_info['Дата окончания проведения ГИКЭ'] = date[:index if index != -1 else len(date)].replace('г', '')
                        if full_time_interval:
                            continue
                    else:
                        date = re.search(r'период с «*\d+»* [А-Яа-яёЁ]+ \d+ г\.*.*[ \n]+по[ \n]+«*\d+»* [А-Яа-яёЁ]+ \d+ г\.*',
                                         text_to_write)
                        if date:
                            current_part += 2
                            text_to_write = re.search(r'по[ \n]+«*\d+»* [А-Яа-яёЁ]+ \d+ г\.*', date.group(0)).group(
                                0).replace('по ', '')
                        year = re.search(r'\d+ г\.', text_to_write)
                        if year:
                            year = year.group(0)
                        elif interval_type != 'words':
                            text_to_write = text
                            year = re.search(r'\d+ г\.', text_to_write).group(0)
                        else:
                            year = re.search(r'\d+ г\.*', text_to_write).group(0)
                        table_info['ГОД'] = year[:year.rfind(' ')]
                        date = re.findall(r'«*\d+»* [А-Яа-яёЁ]+ \d+ г\.*', text_to_write)
                        if date:
                            if len(date) > 1:
                                date = date[1]
                            else:
                                date = date[0]
                        else:
                            date = re.findall(r'«*\d+»*[ \n]+[А-Яа-яёЁ]+[ \n]+\d+ г\.*', text, re.IGNORECASE)[1].replace('  ', ' ')
                        date = date.replace('«', '').replace('»', '')
                        date = date[:date.rfind(' ')]
                        month = re.search(r'[а-яА-ЯёЁ]+', date).group(0)
                        date = date.replace(month, '').replace('  ', '.' + months[month] + '.')
                        day = date[:date.find('.')]
                        if len(day) < 2:
                            date = '0' + date
                        table_info['Дата окончания проведения ГИКЭ'] = date
                        if full_time_interval:
                            continue
                elif act_parts[current_part] == r'\d*\.*.*Место проведения [экспертизы]*:*':
                    if not text_to_write.strip() or re.search(r'«*\d+»*[ \n]+[А-Яа-яёЁ]+[ \n]+\d+ г\.* *\n.[^0-9]+\n', text_to_write, re.IGNORECASE):
                        text_to_write = re.search(r'«*\d+»*[ \n]+[А-Яа-яёЁ]+[ \n]+\d+ г\.* *\n.[^0-9]+\n', text, re.IGNORECASE).group(0)
                        text_to_write = re.search(r'\n.[^0-9]+?(?=\n)', text_to_write, re.IGNORECASE).group(0)
                        broken_structure = True
                    table_info['Место проведения экспертизы'] = text_to_write.replace('–', '').replace(':', '').replace('\n', '')
                elif act_parts[current_part] == r'\d*\.*.*Заказчик экспертизы':
                    if broken_structure:
                        text_to_write = re.search(table_info['Место проведения экспертизы'][:-1] + r'[А-Яа-яёЁ \n,.0-9:"/()«»\\]+', text, re.IGNORECASE)
                        if text_to_write:
                            text_to_write = text_to_write.group(0)[len(table_info['Место проведения экспертизы']):]
                        else:
                            text_to_write = ''
                    table_info['Заказчик работ (*если не указан, то заказчик экспертизы)'] = text_to_write.replace('–', '').replace(':', '')
                elif act_parts[current_part] == r'\d*\.*.*[Сведения об]* эксперте':
                    if re.search(r'Эксперты,[ \n]+состоящие[ \n]+в[ \n]+трудовых', text_to_write, re.IGNORECASE) or several_experts:
                        names = []
                        if not full_name:
                            names = re.findall(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-Яа-яёЁ]+\.[ \n]*[А-Яа-яёЁ]+\.[ \n]+-*–*[ \n]*образование', text_to_write)
                        if not names or several_experts and full_name:
                            names = re.findall(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+[ \n]+-*–*[ \n]*образование',
                                               text_to_write)
                            full_name = True
                        if names:
                            names = list(map(lambda x: x.replace('\n', '').replace(' образование', '').replace(' -', '').replace(' –', ''), names))
                            if several_experts:
                                table_info['Эксперт (физ. или юр.лицо)'] += ',\n' + ',\n'.join(names)
                            else:
                                table_info['Эксперт (физ. или юр.лицо)'] = ',\n'.join(names)
                        several_experts = True
                    else:
                        name = re.search(r'Фамилия,[ \n]*имя[,и \n]*отчество.*[эксперта]*.*\n.*\n', text_to_write, re.IGNORECASE)
                        if name:
                            name = name.group(0)
                            if 'Образование' in name and 'высшее' not in name:
                                name = re.search(r'[А-Яа-яёЁ]+[ \n]*[А-Яа-яёЁ]+[ \n]*[А-Яа-яёЁ]+[ \n]*?(?=\nвысшее)',
                                          text_to_write, re.IGNORECASE).group(0)
                                broken_structure = True
                            else:
                                name = name[re.search(r'Фамилия,[ \n]*имя[,и \n]*отчество.?[эксперта]*:*', name, re.IGNORECASE).end():].replace('\n', '')
                                if 'образование' in name.lower():
                                    find_name = re.search(r'[А-Яа-яёЁ]+[ \n]+[А-Яа-яёЁ]+[ \n]+[А-Яа-яёЁ]+[ \n]*?(?=;)',
                                                     name, re.IGNORECASE)
                                    if not find_name:
                                        name = re.search(r'[А-Яа-яёЁ]+[ \n]+[А-Яа-яёЁ]+[ \n]+[А-Яа-яёЁ]+[ \n]*?(?=Образование)', name, re.IGNORECASE).group(0)
                                    else:
                                        name = find_name.group(0)
                            table_info['Эксперт (физ. или юр.лицо)'] = name
                        else:
                            name = re.search(r'ФИО эксперта.*\n.*\n', text_to_write, re.IGNORECASE)
                            if name:
                                name = name.group(0)
                                name = name[re.search(r'ФИО эксперта.*\n', name, re.IGNORECASE).end():].replace('\n', '')
                                table_info['Эксперт (физ. или юр.лицо)'] = name
                            else:
                                name = re.search(r'[А-Яа-яёЁ]+[ \n]*[А-Яа-яёЁ]+[ \n]*[А-Яа-яёЁ]+[ \n]*?(?=, образование)', text_to_write, re.IGNORECASE)
                                if name:
                                    name = name.group(0)
                                    table_info['Эксперт (физ. или юр.лицо)'] = name
                                else:
                                    with pdfplumber.open(pdf_file) as pdf:
                                        page_tables = pdf.pages[page_number].extract_tables()
                                        if page_tables and len(page_tables[0]) >= 5 and page_tables[0][4][
                                            0] == 'ФИО эксперта':
                                            table_info['Эксперт (физ. или юр.лицо)'] = page_tables[0][4][1]
                elif act_parts[current_part] == 'Объект .*?[экспертизы]*:*':
                    if re.search(r'земли', text_to_write, re.IGNORECASE):
                        table_info['Вид ГИКЭ'] = 'ЗУ'
                    elif re.search(r'раздел', text_to_write, re.IGNORECASE):
                        table_info['Вид ГИКЭ'] = 'НПД'
                    elif re.search(r'документация', text_to_write, re.IGNORECASE):
                        table_info['Вид ГИКЭ'] = 'Док-я'
                    get_gike_object_size(text_to_write, table_info)
                    exp_object = re.search(r'«[/\\А-Яа-яёЁa-zA-Z \n,.0-9:«»-–-()#№+]+?(?=\n\d\.)', text_to_write, re.IGNORECASE)  # «[А-Яа-я \n,.0-9:]+»
                    if not exp_object:
                        exp_object = re.search(r'«[/\\А-Яа-яёЁa-zA-Z \n,.0-9:«»-–-()#№+]+', text_to_write,
                                               re.IGNORECASE)  # «[А-Яа-я \n,.0-9:]+»
                    if exp_object:
                        table_info['Номер (если имеется) и наименование Акта ГИКЭ'] += ' ' + exp_object.group(0)
                        exploration_object = True
                elif act_parts[current_part] == 'Факты и сведения, выявленные .*\n*.*исследований' and not exploration_object:
                    exp_object = re.search(r'«[/\\А-Яа-яёЁa-zA-Z \n,.0-9:«»-–-()#№+]+?(?=Краткая[ \n]+физико-географическая)', text_to_write,
                                           re.IGNORECASE)
                    if exp_object:
                        table_info['Номер (если имеется) и наименование Акта ГИКЭ'] += ' ' + exp_object.group(0)
                elif act_parts[current_part] == 'Вывод экспертизы' and \
                        'Площадь, протяжённость и/или др. параменты объекта' not in table_info.keys():
                    get_gike_object_size(text_to_write, table_info)
                if act_parts[current_part] == 'Сведения о проведенных исследованиях':
                    get_gike_object_size(text_to_write, table_info)
                if act_parts[current_part] == 'Факты и сведения, выявленные .*\n*.*исследований':
                    # object_char = re.search(r'Характеристика[ \n]+объекта', text, re.IGNORECASE)
                    # if object_char:
                    get_gike_object_size(text_to_write, table_info)
                    sectors = re.search(r'Участок[ \n]+№\d+', text, re.IGNORECASE)
                    if sectors and 'Площадь, протяжённость и/или др. параменты объекта' in table_info.keys():
                        sectors = re.findall(r'Участок[ \n]+№\d+[А-Яа-яёЁA-Za-z \n,.0-9:;"/()«»\\–-]+Документация', text, re.IGNORECASE)
                        for sector in sectors:
                            sector = re.search(r'площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*', sector, re.IGNORECASE)
                            # Общ.+[ \n]+площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*
                            if sector:
                                sector = re.search(r'\d+[,]*\d*', sector.group(0),
                                                   re.IGNORECASE).group(0)
                                sectors_square.append(sector)
                        total_square = float(
                            re.search(r'\d+[,]*\d*', table_info['Площадь, протяжённость и/или др. параменты объекта'], re.IGNORECASE).group(0).replace(',', '.'))
                        if total_square and math.isclose(total_square, sum([float(i.replace(',', '.')) for i in sectors_square])):
                            sectors_len = len(sectors_square)
                            table_info['Площадь, протяжённость и/или др. параменты объекта'] += ': всего ' + str(sectors_len) + ' уч-в - '
                            for i in range(sectors_len):
                                if i < sectors_len - 1:
                                    table_info['Площадь, протяжённость и/или др. параменты объекта'] += str(sectors_square[i]).replace('.', ',') + ' + '
                                else:
                                    table_info['Площадь, протяжённость и/или др. параменты объекта'] += str(sectors_square[i]).replace('.', ',')
                                    if 'га' in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                                        table_info['Площадь, протяжённость и/или др. параменты объекта'] += ' га'
                                    elif 'кв. м' in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                                        table_info[
                                            'Площадь, протяжённость и/или др. параменты объекта'] += 'кв. м'
                    if 'Площадь, протяжённость и/или др. параменты объекта' in table_info.keys():
                        perspective = re.search(r'перспект[А-Яа-яёЁA-Za-z \n,.0-9:;"()«»\\–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*', text, re.IGNORECASE)
                        if perspective:
                            enter_reserve = False
                            non_perspective = None
                            small_perspective = None
                            if text_reserve:
                                text_reserve += text
                                text = text_reserve.replace(
                                    '--- Факты и сведения, выявленные .*\n*.*исследований --- ', '')
                                text_reserve = None
                                enter_reserve = True
                            else:
                                non_perspective = re.search(
                                    r'неперспект[А-Яа-яёЁA-Za-z \n,.0-9:;"()«»\\–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*',
                                    text, re.IGNORECASE)
                                small_perspective = re.search(
                                    r'малоперспект[А-Яа-яёЁA-Za-z \n,.0-9:;"()«»\\–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*',
                                    text, re.IGNORECASE)
                                if non_perspective and not small_perspective or not non_perspective and small_perspective:
                                    text_reserve = text
                            if non_perspective and small_perspective or enter_reserve:
                                if non_perspective and 'неперспект' not in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                                    non_perspective = re.search(
                                        r'неперспект[А-Яа-яёЁA-Za-z \n,.0-9:;"()«»\\–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*',
                                        text, re.IGNORECASE)
                                    non_perspective = re.search(r'\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+', non_perspective.group(0), re.IGNORECASE).group(0)
                                    if 'из них' not in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                                        table_info['Площадь, протяжённость и/или др. параменты объекта'] += ' (из них к неперспект. отнесено ' + non_perspective + ', '
                                    else:
                                        table_info['Площадь, протяжённость и/или др. параменты объекта'] += ' к неперспект. - ' + non_perspective + ')'

                                if small_perspective and 'малоперсп' not in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                                    small_perspective = re.search(
                                        r'малоперспект[А-Яа-яёЁA-Za-z \n,.0-9:;"()«»\\–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*',
                                        text, re.IGNORECASE)
                                    small_perspective = re.search(r'\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+',
                                                                small_perspective.group(0), re.IGNORECASE).group(0)
                                    if 'из них' not in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                                        table_info['Площадь, протяжённость и/или др. параменты объекта'] += ' (из них к малоперсп. отнесено ' + small_perspective + ', '
                                    else:
                                        table_info['Площадь, протяжённость и/или др. параменты объекта'] += ' к малоперсп. - ' + small_perspective + ')'
                        if 'из них' not in table_info['Площадь, протяжённость и/или др. параменты объекта']:
                            enter_reserve = False
                            square_object = None
                            line_object = None
                            if text_reserve:
                                text_reserve += text
                                text = text_reserve.replace('--- Факты и сведения, выявленные .*\n*.*исследований --- ', '')
                                text_reserve = None
                                enter_reserve = True
                            else:
                                square_object = re.search(r'Площадной[ \n]+объект', text, re.IGNORECASE)
                                line_object = re.search(r'Линейный[ \n]+объект', text, re.IGNORECASE)
                                if square_object and not line_object or not square_object and line_object:
                                    text_reserve = text
                            if enter_reserve or line_object and square_object:
                                square_object = re.search(
                                    r'Площадной[ \n]+объект[А-Яа-яёЁA-Za-z№# \n,.0-9:;"()«»\\/–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*',
                                    text, re.IGNORECASE)
                                line_object = re.search(
                                    r'Линейный[ \n]+объект[А-Яа-яёЁA-Za-z№# \n,.0-9:;"()«»\\/–-]+?площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[га]*[кв. м]*',
                                    text, re.IGNORECASE)
                                if square_object and line_object:
                                    if 'площ.' not in table_info[
                                        'Площадь, протяжённость и/или др. параменты объекта']:
                                        square_object = re.search(
                                            r'площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+',
                                            square_object.group(0), re.IGNORECASE).group(0)
                                        square_object = re.search(r'\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+',
                                                                  square_object, re.IGNORECASE).group(0)
                                        if 'лин.' not in table_info[
                                            'Площадь, протяжённость и/или др. параменты объекта']:
                                            table_info[
                                                'Площадь, протяжённость и/или др. параменты объекта'] += ' (площ. об. =  ' + square_object + '; '
                                        else:
                                            table_info[
                                                'Площадь, протяжённость и/или др. параменты объекта'] += 'площ. об. = ' + square_object + ')'
                                    if 'лин.' not in table_info[
                                        'Площадь, протяжённость и/или др. параменты объекта']:
                                        line_object = re.search(
                                            r'площадь[ \n]*.*[ \n]*\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+',
                                            line_object.group(0), re.IGNORECASE).group(0)
                                        line_object = re.search(r'\d* *\d+[,]*\d*[ \n]+[а-яА-ЯёЁ]+',
                                                                line_object,
                                                                re.IGNORECASE).group(0)
                                        if 'площ.' not in table_info[
                                            'Площадь, протяжённость и/или др. параменты объекта']:
                                            table_info[
                                                'Площадь, протяжённость и/или др. параменты объекта'] += ' (лин. об. = ' + line_object + '; '
                                        else:
                                            table_info[
                                                'Площадь, протяжённость и/или др. параменты объекта'] += 'лин. об. = ' + line_object + ')'

                        table_info['Площадь, протяжённость и/или др. параменты объекта'] = table_info['Площадь, протяжённость и/или др. параменты объекта'].replace('  ', ' ')
                conclusion = re.search(r'\(.[^\n ]+[^ \n]ельное[ \n]*заключение\)', text_to_write, re.IGNORECASE)
                if conclusion:
                    table_info['Заключение. Выявленые объекты.'] = conclusion.group(0).replace('(', '').replace(')', '')
                else:
                    conclusion = re.search(r'Заключение[ \n]*экспертизы[ \n]*.+', text_to_write, re.IGNORECASE)
                    if conclusion:
                        table_info['Заключение. Выявленые объекты.'] = conclusion.group(0).replace('.', '').replace('Заключение экспертизы ','') + ' заключение'
                        if voan_reserve and 'ВОАН' not in table_info['Заключение. Выявленые объекты.'] and \
                                'отрицательное' in table_info['Заключение. Выявленые объекты.'].lower():
                            table_info['Заключение. Выявленые объекты.'] += voan_reserve

                if not 'ОЛ' in table_info.keys() or 'от' not in table_info['ОЛ'] or '№' not in table_info['ОЛ'] or not re.search(r'[А-ЯЁ]+[а-яё]+', table_info['ОЛ']):
                    #  Открытого[ \n]*листа[ \n]*[а-яА-Я \n0-9.]*№[\d -]+[а-яА-Я \n\d\.,(]*
                    #  Открытого листа[а-яА-Я \n]*№[\d -]+ от [\d. г]+[а-яА-Я \n,(]*на имя.+\..+\. [а-яА-Я]+
                    #  открытого[ \n]*листа[ \n]*[а-яА-Я \n]*№[\d -]+[а-яА-Я \n\d\.,(]*[а-яА-Я]+.+\..+\.

                    open_list = re.search(r'[А-Яа-яёЁ]+\.*[ \n]*[А-Яа-яёЁ]+\.*[ \n]+[А-ЯЁ]+[а-яё]+[ \n]*.*[ \n]*Открыт.*[ \n]*лист.*[ \n]*[а-яА-ЯёЁ \n0-9.]*№[\d -]+[а-яА-ЯёЁ \n\d.,(-«»]*', text_to_write, re.IGNORECASE)
                    if not open_list:
                        open_list = re.search(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-Яа-яёЁ]+\.*[ \n]*[А-Яа-яёЁ]+\.*[ \n]*.*[ \n]*Открыт.*[ \n]*лист.*[ \n]*[а-яА-ЯёЁ \n0-9.]*№[\d -]+[а-яА-ЯёЁ \n\d.,(-«»]*', text_to_write, re.IGNORECASE)
                    if not open_list:
                        open_list = re.search(r'Открытый[ \n]*лист[ \n]*[а-яА-ЯёЁ \n0-9.]*№[\d -]+[а-яА-ЯёЁ \n\d.,(-«»]*?(?=Прил)', text_to_write, re.IGNORECASE)
                    if not open_list:
                        open_list = re.search(r'Открыт.*[ \n]*лист.*[ \n]*[а-яА-ЯёЁ \n0-9.]*№[\d -]+[а-яА-ЯёЁ \n\d.,(-«»]*', text_to_write, re.IGNORECASE)
                    if open_list:
                        open_list = open_list.group(0)
                        list_holder = re.search(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-Яа-яёЁ]+\.[ \n]*[А-Яа-яёЁ]+\.', open_list)
                        if list_holder:
                            list_holder = list_holder.group(0)
                        if not list_holder:
                            list_holder = re.search(r'[А-Яа-яёЁ]+\.[ \n]*[А-Яа-яёЁ]+\.[ \n]+[А-ЯЁ]+[а-яё]+', open_list)
                            if list_holder:
                                list_holder = list_holder.group(0)
                        if not list_holder:
                            list_holder = re.search(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+', open_list)
                            if list_holder:
                                list_holder = list_holder.group(0)
                                ''' Замена ИО на инициалы (иногда имя пишут первым, поэтому использование не всегда оправдано)
                                capital_words = re.findall(r'[А-ЯЁ]+', list_holder)
                                list_holder = list_holder[:list_holder.find(' ')]
                                list_holder += ' ' + capital_words[1] + '.' + capital_words[2] + '.'
                                '''
                        if not list_holder:
                            list_holder = ''
                        list_number = re.search(r'№[ \n]*\d+-\d+', open_list, re.IGNORECASE)
                        if list_number:
                            list_number = list_number.group(0)
                        else:
                            list_number = ''
                        list_date = re.search(r'\d{2}.\d{2}.\d{4}', open_list, re.IGNORECASE)
                        if list_date:
                            list_date = list_date.group(0)
                        if not list_date:
                            list_date = re.search(r'«*\d+»* [А-Яа-яёЁ]+ \d{4}', open_list, re.IGNORECASE)
                            if list_date:
                                list_date = list_date.group(0).replace('«', '').replace('»', '')
                                list_month = re.search(r'[а-яА-ЯёЁ]+', list_date).group(0)
                                list_date = date.replace(list_month, '').replace('  ', '.' + months[list_month] + '.')
                                list_day = date[:date.find('.')]
                                if len(list_day) < 2:
                                    list_date = '0' + list_date
                        if list_date:
                            list_date = ' от ' + list_date
                        else:
                            list_date = ''
                        table_info['ОЛ'] = list_holder + list_date + ' ' + list_number

                voan = re.search(r'выявлен[\n ]+объект[\n ]+археологического[\n ]+наследия[\n ]+.*«.*»', text, re.IGNORECASE)
                if not voan:
                    voan = re.search(r'выявлен[\n ]+объект[\n ]+археологического[\n ]+наследия[\n ]+.*".*"', text,re.IGNORECASE)
                if not voan:
                    voan = re.search(r'ВОАН[\n ]+.*«.*»', text,re.IGNORECASE)
                if voan:
                    voan = voan.group(0)
                    voan = ' ВОАН ' + voan[voan.find('«')-1:]
                    if 'Заключение. Выявленые объекты.' in table_info.keys() and 'ВОАН' not in table_info['Заключение. Выявленые объекты.'] and \
                            'отрицательное' in table_info['Заключение. Выявленые объекты.'].lower():
                        table_info['Заключение. Выявленые объекты.'] += voan
                    else:
                        voan_reserve = voan

                annotation = re.search(r'аннотация', text, re.IGNORECASE)
                if annotation:
                    get_gike_object_size(text_to_write, table_info)

                executor = re.search(r'Директор [а-яА-ЯёЁa-zA-Z\n«»" -]+.{1}\..{1}\..+', text, re.IGNORECASE)
                if executor and re.search(r'Эксперт', text, re.IGNORECASE):
                    executor = executor.group(0).replace('Директор ', '').replace('директор ', '')
                    executor = executor[:re.search(r'.{1}\..{1}\..+', executor, re.IGNORECASE).start()]
                    table_info['Исполнитель полевых работ (юр. лицо)'] = executor
                else:
                    executor = re.search(r'Полное[ \n]*и[ \n]*сокращенное[ \n]*наименование[ \n]*организации[а-яА-ЯёЁa-zA-Z\n«»" -()]+?(?=Организационно)', text,
                              re.IGNORECASE)
                    if executor:
                        res = executor.group(0)
                        executor = res[re.search(r'Полное[ \n]*и[ \n]*сокращенное[ \n]*наименование[ \n]*организации[ \n]*', res, re.IGNORECASE).end():]
                        table_info['Исполнитель полевых работ (юр. лицо)'] = executor

                if next_index:
                    current_part += 1
                    continue
                elif current_part + 2 < len(act_parts) and re.search(act_parts[current_part+2].replace(' ', r'[ \n]*'), text):
                    current_part += 2
                    continue
                break

    if ('Площадь, протяжённость и/или др. параменты объекта' not in table_info.keys() or \
            'Общ. S' not in table_info['Площадь, протяжённость и/или др. параменты объекта']) and len(SQUARE_RESERVE) > 0:
        table_info['Площадь, протяжённость и/или др. параменты объекта'] = 'Общ. S = ' + SQUARE_RESERVE[0]
    document.close()

    # pd.DataFrame(table_info,columns=table_columns,index=[0]).to_excel(folder + "/" + "table.xlsx", index=False, engine='openpyxl')
    table_path = "uploaded_files/РЕЕСТР актов ГИКЭ.xlsx"
    df_new = pd.DataFrame(table_info, columns=table_columns, index=[0])
    table_data = df_new
    if os.path.exists(table_path):
        df_existing = pd.read_excel(table_path)
        df_new = df_existing._append(df_new, ignore_index=True)
    df_str = df_new.astype(pd.StringDtype())
    df_str = df_str.map(lambda x: x if re.search(r'[А-Яа-яёЁA-Za-z.0-9,]+', str(x), re.IGNORECASE) else np.nan)
    cells_sum = df_str.size - df_str.isnull().sum().sum()
    print(str(cells_sum) + '/' + str(df_str.size), str(round(cells_sum / df_str.size * 100, 2)) + '%')
    df_new['Дата окончания проведения ГИКЭ'] = pd.to_datetime(df_new['Дата окончания проведения ГИКЭ'], format='%d.%m.%Y', dayfirst=True)
    # df_new['Дата окончания проведения ГИКЭ'] = df_new['Дата окончания проведения ГИКЭ'].dt.date
    # df_new['Дата окончания проведения ГИКЭ'] = df_new['Дата окончания проведения ГИКЭ'] = [x.strftime("%d-%m-%y") for x in df_new.date]
    df_new.sort_values(by='Дата окончания проведения ГИКЭ', ascending=False, inplace=True)
    df_new['Дата окончания проведения ГИКЭ'] = df_new['Дата окончания проведения ГИКЭ'].dt.strftime('%d.%m.%Y')
    with pd.ExcelWriter(table_path) as writer:
        df_new.to_excel(writer, sheet_name="Sheet1", index=False)
        # auto_adjust_xlsx_column_width(df_new, writer, sheet_name="Sheet1", margin=10)
    wb = load_workbook(table_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 6.86
    ws.column_dimensions['B'].width = 10.14
    ws.column_dimensions['C'].width = 10.14
    ws.column_dimensions['D'].width = 66.43
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 26
    ws.column_dimensions['G'].width = 20.71
    ws.column_dimensions['H'].width = 18.43
    ws.column_dimensions['I'].width = 24.71
    ws.column_dimensions['J'].width = 21.29
    ws.column_dimensions['K'].width = 26
    ws.column_dimensions['L'].width = 27.29
    font = Font(
        name='Times New Roman',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}
    for i in range(1, len(df_new.values)+2):
        if i == 1:
            ws.row_dimensions[0].height = 50
        else:
            ws.row_dimensions[i].height = 80
        for cell in ws[i]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(table_path)
    return table_data
