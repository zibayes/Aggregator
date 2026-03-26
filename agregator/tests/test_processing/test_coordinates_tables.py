import pytest
import pandas as pd
import numpy as np
import openpyxl
from unittest.mock import patch, MagicMock
from agregator.processing.coordinates_tables import (
    analyze_coordinates_in_tables_from_pdf,
    check_table_numbers_join,
    extract_tables_from_docx,
    extract_coordinates_from_docx_table,
    extract_coordinates_xlsx,
    search_coords_in_text,
    check_tables_joining,
    check_is_coordinate_table,
    format_coordinates,
    extract_text_from_pdf,
    extract_tables_from_pdf,
    fill_dataframe_from_pdf,
)
from agregator.processing.geo_utils import COORDINATE_SYSTEMS, COORDINATE_MARKS


# ---- Тесты для вспомогательных функций ----
def test_check_table_numbers_join():
    df = pd.DataFrame({'col1': ['1', '2', '3', '4']})
    assert check_table_numbers_join(0, 0, '1', df, 4, 1) is True

    df2 = pd.DataFrame({'col1': ['1', '5', '6']})
    assert check_table_numbers_join(0, 0, '1', df2, 3, 1) is True
    assert check_table_numbers_join(0, 0, '5', df2, 3, 1) is True  # 5 <= 5 <= 7 -> True
    assert check_table_numbers_join(0, 0, '10', df2, 3, 1) is False  # 10 не в диапазоне


def test_check_tables_joining():
    table1 = [['1', 'X', 'Y'], ['2', 'X', 'Y']]
    table2 = [['3', 'X', 'Y']]
    assert check_tables_joining(table1, table2) is True

    # Не подходит по столбцам
    table2_diff = [['3', 'X']]
    assert check_tables_joining(table1, table2_diff) is False

    # Не координатная таблица (два столбца)
    table_not_coord = [['A', 'B'], ['1', '2']]  # два столбца, чтобы не было IndexError
    assert check_tables_joining(table1, table_not_coord) is False


def test_check_is_coordinate_table():
    coord_table = [
        ['№', 'Северная широта', 'Восточная долгота'],
        ['1', '55°45\'30"', '92°30\'15"']
    ]
    assert check_is_coordinate_table(coord_table) is True

    coord_table2 = [
        ['№', 'X', 'Y'],
        ['1', '123.45', '456.78']
    ]
    assert check_is_coordinate_table(coord_table2) is True

    # Не координатная таблица (нет чисел во втором столбце)
    not_coord = [['А', 'Б'], ['в', 'г']]  # нет чисел -> false
    assert check_is_coordinate_table(not_coord) is False


# ---- Тесты для extract_text_from_pdf ----
@patch('agregator.processing.coordinates_tables.pdfplumber')
def test_extract_text_from_pdf_success(mock_pdfplumber):
    mock_pdf = MagicMock()
    mock_page = MagicMock()
    mock_page.extract_text.return_value = "Текст страницы"
    mock_pdf.pages = [mock_page]
    mock_pdfplumber.open.return_value.__enter__.return_value = mock_pdf

    result = extract_text_from_pdf("test.pdf")
    assert result == "Текст страницы"


@patch('agregator.processing.coordinates_tables.pdfplumber')
def test_extract_text_from_pdf_failure(mock_pdfplumber):
    mock_pdfplumber.open.side_effect = Exception("Ошибка")
    result = extract_text_from_pdf("test.pdf")
    assert result is None


# ---- Тесты для extract_tables_from_pdf ----
@patch('agregator.processing.coordinates_tables.pdfplumber')
def test_extract_tables_from_pdf_success(mock_pdfplumber):
    mock_pdf = MagicMock()
    mock_page = MagicMock()
    mock_table = [['A', 'B'], ['1', '2']]
    mock_page.extract_tables.return_value = [mock_table]
    mock_pdf.pages = [mock_page]
    mock_pdfplumber.open.return_value.__enter__.return_value = mock_pdf

    result = extract_tables_from_pdf("test.pdf")
    assert result == [mock_table]


@patch('agregator.processing.coordinates_tables.pdfplumber')
def test_extract_tables_from_pdf_failure(mock_pdfplumber):
    mock_pdfplumber.open.side_effect = Exception("Ошибка")
    result = extract_tables_from_pdf("test.pdf")
    assert result is None


# ---- Тесты для extract_tables_from_docx ----
@patch('agregator.processing.coordinates_tables.Document')
def test_extract_tables_from_docx(mock_doc):
    mock_document = MagicMock()
    mock_table = MagicMock()
    mock_row = MagicMock()
    mock_cell = MagicMock()
    mock_cell.text = "Значение"
    mock_row.cells = [mock_cell]
    mock_table.rows = [mock_row]
    mock_document.tables = [mock_table]
    mock_doc.return_value = mock_document

    result = extract_tables_from_docx("test.docx")
    assert len(result) == 1
    df = result[0]
    assert df.iloc[0, 0] == "Значение"


# ---- Тесты для extract_coordinates_from_docx_table ----
@patch('agregator.processing.coordinates_tables.COORDINATE_SYSTEMS', ['wgs84'])
@patch('agregator.processing.coordinates_tables.COORDINATE_MARKS', {('север', 'вост'): False})
def test_extract_coordinates_from_docx_table():
    # Создаем мок документа с таблицей
    doc = MagicMock()
    doc.paragraphs = [MagicMock(text="wgs84")]

    # Создаем мок таблицы
    mock_table = MagicMock()
    mock_row1 = MagicMock()
    mock_cell1_1 = MagicMock()
    mock_cell1_1.text = "Северная широта"
    mock_cell1_2 = MagicMock()
    mock_cell1_2.text = "Восточная долгота"
    mock_row1.cells = [mock_cell1_1, mock_cell1_2]

    mock_row2 = MagicMock()
    mock_cell2_1 = MagicMock()
    mock_cell2_1.text = "55°45'30\""
    mock_cell2_2 = MagicMock()
    mock_cell2_2.text = "92°30'15\""
    mock_row2.cells = [mock_cell2_1, mock_cell2_2]

    mock_table.rows = [mock_row1, mock_row2]

    # Запускаем
    result = extract_coordinates_from_docx_table(mock_table, doc)
    assert result is not None
    dfs, coord_sys = result
    assert len(dfs) > 0
    assert 'wgs84' in coord_sys


# ---- Тесты для extract_coordinates_xlsx ----
def test_extract_coordinates_xlsx(tmp_path):
    # Создаем временный xlsx файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Северная широта'
    ws['B1'] = 'Восточная долгота'
    ws['A2'] = '55°45\'30"'
    ws['B2'] = '92°30\'15"'
    ws['C1'] = 'wgs84'  # система координат
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)

    with patch('agregator.processing.coordinates_tables.COORDINATE_SYSTEMS', ['wgs84']):
        result = extract_coordinates_xlsx(str(file_path))
        assert result is not None
        dfs, coord_sys = result
        assert len(dfs) > 0
        assert 'wgs84' in coord_sys


# ---- Тесты для search_coords_in_text ----
@patch('agregator.processing.coordinates_tables.calculate_polygons_area')
@patch('agregator.processing.coordinates_tables.check_is_coordinate_table')
@patch('agregator.processing.coordinates_tables.check_tables_joining')
def test_search_coords_in_text_no_tables(mock_join, mock_is_coord, mock_calc):
    pdf = MagicMock()
    pdf.pages = [MagicMock(extract_tables=lambda: [])]
    document = [MagicMock(get_text=lambda: "Текст без координат")]
    tables = []
    coordinates = {}

    search_coords_in_text(pdf, 0, document, tables, "Текст без координат", coordinates)

    # Ничего не должно измениться
    assert coordinates == {}
    assert tables == []


# ---- Тесты для format_coordinates ----
def test_format_coordinates():
    # Добавляем десятичную часть в секунды, чтобы попасть под regex
    df = pd.DataFrame({
        'A': ['№', '1', '2'],
        'B': ['X', '55°45\'30.0"', '55°46\'00.0"'],
        'C': ['Y', '92°30\'15.0"', '92°31\'00.0"']
    })
    results = [df]
    coordinate_systems = ['wgs84']

    result = format_coordinates(results, coordinate_systems)
    # Проверяем, что ключ содержит "Каталог координат (wgs84)" и что есть три точки
    assert 'Каталог координат (wgs84)' in result
    coords = result['Каталог координат (wgs84)']
    assert coords['coordinate_system'] == 'wgs84'
    assert len(coords) == 3  # две точки + 'coordinate_system'
    assert 1 in coords
    assert 2 in coords


def test_format_coordinates_with_zone():
    df = pd.DataFrame({
        'A': ['Участок', 'Участок 1', '', ''],
        'B': ['№', '1', '2', '3'],
        'C': ['X', '55°45\'30.0"', '55°46\'00.0"', '55°46\'30.0"'],
        'D': ['Y', '92°30\'15.0"', '92°31\'00.0"', '92°31\'30.0"']
    })
    results = [df]
    coordinate_systems = ['wgs84']

    result = format_coordinates(results, coordinate_systems)
    # Находим ключ, который начинается с "Участок 1" и содержит (wgs84)
    matching_keys = [k for k in result.keys() if k.startswith('Участок 1')]
    assert len(matching_keys) > 0
    # Берём ключ с максимальным суффиксом (он должен содержать все точки)
    main_key = max(matching_keys, key=lambda k: len(k))
    coords = result[main_key]
    assert coords['coordinate_system'] == 'wgs84'
    assert len(coords) == 2


# ---- Тесты для fill_dataframe_from_pdf ----
def test_fill_dataframe_from_pdf():
    class CellMock:
        def __init__(self, row, col):
            self.row = row
            self.column = col

    df = pd.DataFrame({
        'A': ['Северная широта', '55°45\'30"', '55°46\'00"'],
        'B': ['Восточная долгота', '92°30\'15"', '92°31\'00"'],
        'C': ['Доп', 'x', 'y']
    })
    target_cell = CellMock(2, 2)  # ячейка с данными (вторая строка, второй столбец)
    dfs = []
    multiple_coord_sys = False
    coordinate_systems = ['wgs84']
    current_area = [None]
    legend_length = 2

    fill_dataframe_from_pdf(target_cell, df, dfs, multiple_coord_sys, coordinate_systems, current_area, legend_length)
    assert len(dfs) == 1
    # Функция выбирает диапазон столбцов [0,1] – два столбца
    assert dfs[0].shape == (2, 2)


# ---- Тесты для analyze_coordinates_in_tables_from_pdf (частичные) ----
@patch('agregator.processing.coordinates_tables.extract_text_from_pdf')
def test_analyze_coordinates_in_tables_from_pdf_no_tables(mock_extract):
    mock_extract.return_value = "Текст"
    tables = []
    result = analyze_coordinates_in_tables_from_pdf(tables, "test.pdf")
    assert result == (None, None, None)


@patch('agregator.processing.coordinates_tables.extract_text_from_pdf')
@patch('agregator.processing.coordinates_tables.fill_dataframe_from_pdf')
def test_analyze_coordinates_in_tables_from_pdf_with_coords(mock_fill, mock_extract):
    mock_extract.return_value = "wgs84"
    tables = [[['Северная широта', 'Восточная долгота'], ['55°45\'30"', '92°30\'15"']]]
    result = analyze_coordinates_in_tables_from_pdf(tables, "test.pdf")
    # Проверяем, что вернулись dfs
    assert result[0] is not None
    assert result[1] is not None
    assert result[2] is not None


# ---- Запуск тестов ----
if __name__ == "__main__":
    pytest.main([__file__, "-v"])
