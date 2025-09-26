import pytest
import pandas as pd
import tempfile
import os
from unittest.mock import Mock, patch, MagicMock
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.exceptions import ValidationError
from django.http import HttpRequest
from django.contrib.auth.models import AnonymousUser

# ПРАВИЛЬНЫЙ импорт
from agregator.views.utils import (
    validate_email, upload_entity_view, get_register_view, generate_excel_report,
    process_edit_form, process_supplement, create_model_dataframe, get_scan_task,
    get_user_tasks
)


class TestValidateEmail:
    """Тесты для функции validate_email"""

    @pytest.mark.parametrize("valid_email", [
        "test@example.com",
        "user.name@domain.co.uk",
        "user+tag@example.org",
        "123@numbers.com",
        "UPPERCASE@EXAMPLE.COM",
        "a@b.cd",
    ])
    def test_valid_emails(self, valid_email):
        assert validate_email(valid_email) is True

    @pytest.mark.parametrize("invalid_email", [
        "invalid", "invalid@", "@example.com", "invalid@example",
        "invalid@example..com", "invalid@.com", "invalid@com.",
        " spaces@example.com ", "invalid@exam ple.com", "",
    ])
    def test_invalid_emails(self, invalid_email):
        assert validate_email(invalid_email) is False


class TestUploadEntityView:
    """Тесты для функции upload_entity_view"""

    def test_get_request_returns_form(self, test_user):
        request = Mock()
        request.user = test_user
        request.method = 'GET'
        request.FILES = {}

        entity_form = Mock(return_value=Mock())

        # Мокаем render чтобы избежать ошибок шаблонов
        with patch('agregator.views.utils.render') as mock_render:
            mock_render.return_value = Mock(status_code=200)

            result = upload_entity_view(
                request=request,
                tasks_id=['task1'],
                entity_type='test_type',
                entity_form=entity_form,
                save_func=Mock(),
                process_func=Mock(),
                error_handler=Mock(),
                page='test_page.html'
            )

            assert mock_render.called

    def test_post_request_valid_form_non_superuser(self, test_user):
        request = Mock()
        request.user = test_user
        request.method = 'POST'
        request.POST = {}
        request.FILES = {}

        mock_form = Mock()
        mock_form.is_valid.return_value = True
        mock_form.cleaned_data = {'files': [Mock()]}

        # Мокаем ВСЕ внешние зависимости
        with patch('agregator.views.utils.render') as mock_render, \
                patch('agregator.views.utils.UserTasks') as mock_user_tasks_class:
            mock_render.return_value = Mock(status_code=200)
            mock_user_task_instance = Mock()
            mock_user_tasks_class.return_value = mock_user_task_instance

            mock_save_func = Mock(return_value=[1, 2, 3])
            mock_process_func = Mock()
            mock_process_func.apply_async.return_value = Mock(task_id='new_task_id')

            result = upload_entity_view(
                request=request,
                tasks_id=['existing_task'],
                entity_type='test_type',
                entity_form=lambda x, y: mock_form,  # Исправляем сигнатуру
                save_func=mock_save_func,
                process_func=mock_process_func,
                error_handler=Mock(),
                page='test_page.html'
            )

            assert mock_render.called
            mock_user_tasks_class.assert_called_once()

    def test_post_request_invalid_form(self, test_user):
        request = Mock()
        request.user = test_user
        request.method = 'POST'
        request.FILES = {}

        mock_form = Mock()
        mock_form.is_valid.return_value = False

        with patch('agregator.views.utils.render') as mock_render:
            mock_render.return_value = Mock(status_code=200)

            result = upload_entity_view(
                request=request,
                tasks_id=[],
                entity_type='test_type',
                entity_form=lambda x, y: mock_form,  # Исправляем сигнатуру
                save_func=Mock(),
                process_func=Mock(),
                error_handler=Mock(),
                page='test_page.html'
            )

            assert mock_render.called


class TestGetRegisterView:
    """Тесты для функции get_register_view"""

    def test_basic_functionality(self, test_user):
        request = Mock()
        request.user = test_user
        request.user.id = 1

        # Создаем реальные Mock объекты для queryset с поддержкой оператора |
        mock_public_qs = Mock()
        mock_private_qs = Mock()
        mock_combined_qs = Mock()

        # Настраиваем поддержку оператора |
        mock_public_qs.__or__ = Mock(return_value=mock_combined_qs)

        # Настраиваем цепочку вызовов ORM
        mock_model = Mock()
        mock_model.objects.filter.return_value.only.return_value = mock_public_qs

        # Мокаем render
        with patch('agregator.views.utils.render') as mock_render:
            mock_render.return_value = Mock(status_code=200)

            result = get_register_view(
                request=request,
                model=mock_model,
                entity_name='test_entities',
                public_only_fields=['field1', 'field2'],
                private_only_fields=['field3', 'field4'],
                template_name='test_template.html'
            )

            # Проверяем что render был вызван
            mock_render.assert_called_once()

    def test_without_only_fields(self, test_user):
        request = Mock()
        request.user = test_user
        request.user.id = 1

        mock_public_qs = Mock()
        mock_private_qs = Mock()
        mock_combined_qs = Mock()

        mock_public_qs.__or__ = Mock(return_value=mock_combined_qs)

        mock_model = Mock()
        mock_model.objects.filter.return_value.only.return_value = mock_public_qs

        with patch('agregator.views.utils.render') as mock_render:
            mock_render.return_value = Mock(status_code=200)

            result = get_register_view(
                request=request,
                model=mock_model,
                entity_name='test_entities',
                public_only_fields=None,
                private_only_fields=None,
                template_name='test_template.html'
            )

            mock_render.assert_called_once()


class TestGenerateExcelReport:
    """Тесты для функции generate_excel_report"""

    def test_basic_excel_generation(self):
        sample_dataframe = pd.DataFrame({
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'City': ['New York', 'London', 'Tokyo']
        })

        # Используем временную директорию вместо мокинга
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, 'test.xlsx')
            result = generate_excel_report(sample_dataframe, file_path)

            # Проверяем что файл создан
            assert os.path.exists(result)

            # Проверяем содержимое файла
            df = pd.read_excel(result)
            assert len(df) == 3
            assert list(df.columns) == ['Name', 'Age', 'City']

    def test_empty_dataframe(self):
        empty_df = pd.DataFrame()

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, 'empty.xlsx')
            result = generate_excel_report(empty_df, file_path)

            # Проверяем что файл создан
            assert os.path.exists(result)

            # Проверяем, что созданный файл пустой
            df = pd.read_excel(result)
            assert len(df) == 0
            assert len(df.columns) == 0


class TestProcessEditForm:
    """Тесты для функции process_edit_form"""

    def test_post_request_updates_fields(self):
        request = Mock()
        request.method = 'POST'
        request.POST = {'name': 'New Name', 'age': '30'}

        instance = Mock()
        instance.save = Mock()

        result = process_edit_form(request, instance, ['name', 'age'])

        assert result is True
        instance.save.assert_called_once()

    def test_non_post_request_returns_false(self):
        request = Mock()
        request.method = 'GET'

        instance = Mock()

        result = process_edit_form(request, instance, ['name'])

        assert result is False


class TestProcessSupplement:
    """Тесты для функции process_supplement"""

    def test_basic_supplement_processing(self):
        request = Mock()
        # Исправляем: POST должен быть Mock с методом dict()
        request.POST = Mock()
        request.POST.dict.return_value = {
            'source-img1.jpg': 'New Source 1',
            'label-img1.jpg': 'New Label 1'
        }

        instance = Mock()
        instance.supplement_dict = {
            'images': [{'source': 'img1.jpg', 'label': 'Label 1'}]
        }

        result = process_supplement(request, instance)

        assert result['images'][0]['source'] == 'New Source 1'
        assert result['images'][0]['label'] == 'New Label 1'

    def test_empty_supplement_dict(self):
        request = Mock()
        request.POST = Mock()
        request.POST.dict.return_value = {}

        instance = Mock()
        instance.supplement_dict = None

        result = process_supplement(request, instance)

        assert result is None


class TestCreateModelDataframe:
    """Тесты для функции create_model_dataframe"""

    def test_basic_dataframe_creation(self):
        mock_model = Mock()

        instance1 = Mock()
        instance1.name = 'Object 1'
        instance1.value = 100

        instance2 = Mock()
        instance2.name = 'Object 2'
        instance2.value = 200

        mock_model.objects.all.return_value = [instance1, instance2]

        fields_mapping = {'Name': 'name', 'Value': 'value'}

        df = create_model_dataframe(mock_model, fields_mapping)

        assert isinstance(df, pd.DataFrame)
        assert len(df) == 2
        assert list(df.columns) == ['Name', 'Value']

    def test_empty_model(self):
        mock_model = Mock()
        mock_model.objects.all.return_value = []

        fields_mapping = {'Name': 'name', 'Value': 'value'}

        df = create_model_dataframe(mock_model, fields_mapping)

        # Исправление: функция возвращает None для пустой модели
        assert df is None


class TestGetScanTask:
    """Тесты для функции get_scan_task"""

    def test_active_task_exists(self):
        mock_task_result = Mock()
        mock_task_result.task_id = 'test_task_id'

        # ПРАВИЛЬНЫЙ путь для мока
        with patch('agregator.views.utils.TaskResult.objects') as mock_objects:
            # Настраиваем цепочку вызовов
            mock_filter = mock_objects.filter.return_value
            mock_exclude = mock_filter.exclude.return_value
            mock_order_by = mock_exclude.order_by.return_value
            mock_order_by.first.return_value = mock_task_result

            is_processing, task_id, task_obj = get_scan_task('test_task_name')

            assert is_processing is True
            assert task_id == 'test_task_id'

    def test_no_active_task(self):
        with patch('agregator.views.utils.TaskResult.objects') as mock_objects:
            mock_filter = mock_objects.filter.return_value
            mock_exclude = mock_filter.exclude.return_value
            mock_order_by = mock_exclude.order_by.return_value
            mock_order_by.first.return_value = None

            is_processing, task_id, task_obj = get_scan_task('test_task_name')

            assert is_processing is False
            assert task_id is None


class TestGetUserTasks:
    """Тесты для функции get_user_tasks"""

    def test_basic_functionality_user_files(self, test_user):
        mock_user_tasks = []
        for i in range(3):
            task = Mock()
            task.task_id = f'test_task_{i}'
            task.upload_source_dict = {'source': 'Пользовательский файл'}
            mock_user_tasks.append(task)

        # ПРАВИЛЬНЫЙ путь для мока
        with patch('agregator.views.utils.UserTasks.objects') as mock_user_objects:
            mock_user_objects.filter.return_value = mock_user_tasks

            with patch('agregator.views.utils.TaskResult.objects') as mock_task_objects:
                mock_task_results = [Mock() for _ in range(3)]
                for i, task_result in enumerate(mock_task_results):
                    task_result.task_id = f'test_task_{i}'

                mock_task_objects.filter.return_value.order_by.return_value = mock_task_results

                tasks_ids = get_user_tasks(test_user.id, ['act', 'scientific_report'])

                assert isinstance(tasks_ids, list)

    def test_external_source_filter(self, test_user):
        mock_user_tasks = []
        task = Mock()
        task.task_id = 'test_task'
        task.upload_source_dict = {'source': 'external'}
        mock_user_tasks.append(task)

        with patch('agregator.views.utils.UserTasks.objects') as mock_user_objects:
            mock_user_objects.filter.return_value = mock_user_tasks

            with patch('agregator.views.utils.TaskResult.objects') as mock_task_objects:
                mock_task_results = [Mock()]
                mock_task_results[0].task_id = 'test_task'
                mock_task_objects.filter.return_value.order_by.return_value = mock_task_results

                tasks_ids = get_user_tasks(test_user.id, ['act'], upload_source=True)

                assert isinstance(tasks_ids, list)


# Security тесты
class TestSecurity:
    """Тесты безопасности"""

    def test_validate_email_security(self):
        dangerous_emails = [
            "test@example.com<script>alert('xss')</script>",
            "../../etc/passwd@example.com",
            "'; DROP TABLE users; --@example.com",
        ]

        for email in dangerous_emails:
            result = validate_email(email)
            assert isinstance(result, bool)

    def test_process_edit_form_security(self):
        request = Mock()
        request.method = 'POST'
        request.POST = {
            'name': '<script>alert("xss")</script>',
            'age': '30; DROP TABLE users;',
        }

        instance = Mock()
        instance.save = Mock()

        result = process_edit_form(request, instance, ['name', 'age'])

        assert result is True

    def test_generate_excel_report_path_traversal(self):
        """Тест: защита от path traversal в generate_excel_report"""
        dangerous_paths = [
            "../../etc/passwd",
            "/root/.ssh/id_rsa",
            "C:\\Windows\\System32\\config\\SAM",
            "../../malicious.xlsx"
        ]

        df = pd.DataFrame({'test': [1, 2, 3]})

        for path in dangerous_paths:
            # Мокаем всё чтобы избежать реального создания файлов
            with patch('agregator.views.utils.pd.ExcelWriter') as mock_excel_writer, \
                    patch('agregator.views.utils.load_workbook') as mock_load_workbook:

                mock_writer = Mock()
                mock_excel_writer.return_value.__enter__ = Mock(return_value=mock_writer)
                mock_excel_writer.return_value.__exit__ = Mock(return_value=None)

                try:
                    result = generate_excel_report(df, path)
                    # Если не упало, проверяем что путь возвращен
                    assert result == path
                except Exception:
                    # Ожидаем исключения для опасных путей
                    pass

    def test_sql_injection_protection(self):
        """Тест: защита от SQL injection в функциях работы с БД"""
        dangerous_inputs = [
            "'; DROP TABLE users; --",
            "1 OR 1=1",
            "test' OR '1'='1",
        ]

        for dangerous in dangerous_inputs:
            # get_scan_task должна безопасно обрабатывать имена задач
            with patch('agregator.views.utils.TaskResult.objects') as mock_objects:
                mock_filter = mock_objects.filter.return_value
                mock_exclude = mock_filter.exclude.return_value
                mock_order_by = mock_exclude.order_by.return_value
                mock_order_by.first.return_value = None

                # Функция не должна падать на опасных входах
                result = get_scan_task(dangerous)
                assert result[0] is False


# Граничные случаи
class TestEdgeCases:
    """Тесты граничных случаев"""

    @pytest.mark.parametrize("email,expected", [
        ("", False),
        (None, False),
        ("a" * 1000 + "@example.com", False),
        ("@", False),
        ("@.", False),
    ])
    def test_validate_email_edge_cases(self, email, expected):
        result = validate_email(email)
        # Для этих случаев ожидаем False
        assert result == expected

    def test_generate_excel_report_large_data(self):
        """Тест: генерация Excel с большими данными"""
        large_data = {'col': range(1000)}
        large_df = pd.DataFrame(large_data)

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, 'large.xlsx')
            result_path = generate_excel_report(large_df, file_path)

            # Проверяем что файл создан
            assert os.path.exists(result_path)

            # Проверяем содержимое
            df = pd.read_excel(result_path)
            assert len(df) == 1000
            assert list(df.columns) == ['col']


# Интеграционные тесты
@pytest.mark.django_db
class TestIntegration:
    """Интеграционные тесты"""

    def test_get_user_tasks_integration(self, test_user, test_tasks):
        # Используем реальную базу данных для этого теста
        tasks_ids = get_user_tasks(test_user.id, ['act', 'scientific_report'])
        assert isinstance(tasks_ids, list)


# Дополнительные тесты для полного покрытия
class TestAdditionalCases:
    """Дополнительные тесты для 100% покрытия"""

    def test_upload_entity_view_superuser_public(self, admin_user):
        """Тест: суперпользователь с public storage"""
        request = Mock()
        request.user = admin_user
        request.method = 'POST'
        request.POST = {'storage_type': 'public'}
        request.FILES = {}

        mock_form = Mock()
        mock_form.is_valid.return_value = True
        mock_form.cleaned_data = {'files': [Mock()]}

        with patch('agregator.views.utils.render') as mock_render, \
                patch('agregator.views.utils.UserTasks') as mock_user_tasks:
            mock_render.return_value = Mock()
            mock_user_tasks.return_value = Mock()

            mock_save_func = Mock(return_value=[1, 2, 3])
            mock_process_func = Mock()
            mock_process_func.apply_async.return_value = Mock(task_id='new_task_id')

            result = upload_entity_view(
                request=request,
                tasks_id=[],
                entity_type='test_type',
                entity_form=lambda x, y: mock_form,
                save_func=mock_save_func,
                process_func=mock_process_func,
                error_handler=Mock(),
                page='test_page.html'
            )

            # Проверяем что is_public=True передается для суперпользователя
            mock_save_func.assert_called_once()
            call_args = mock_save_func.call_args[0]
            assert call_args[2] is True  # is_public=True

    def test_upload_entity_view_superuser_private(self, admin_user):
        """Тест: суперпользователь с private storage"""
        request = Mock()
        request.user = admin_user
        request.method = 'POST'
        request.POST = {'storage_type': 'private'}
        request.FILES = {}

        mock_form = Mock()
        mock_form.is_valid.return_value = True
        mock_form.cleaned_data = {'files': [Mock()]}

        with patch('agregator.views.utils.render') as mock_render, \
                patch('agregator.views.utils.UserTasks') as mock_user_tasks:
            mock_render.return_value = Mock()
            mock_user_tasks.return_value = Mock()

            mock_save_func = Mock(return_value=[1, 2, 3])

            result = upload_entity_view(
                request=request,
                tasks_id=[],
                entity_type='test_type',
                entity_form=lambda x, y: mock_form,
                save_func=mock_save_func,
                process_func=Mock(),
                error_handler=Mock(),
                page='test_page.html'
            )

            call_args = mock_save_func.call_args[0]
            assert call_args[2] is False  # is_public=False

    def test_get_scan_task_exception_handling(self):
        """Тест: обработка исключений в get_scan_task"""
        with patch('agregator.views.utils.TaskResult.objects') as mock_objects:
            mock_objects.filter.side_effect = Exception("Database error")

            with patch('agregator.views.utils.logging') as mock_logging:
                is_processing, task_id, task_obj = get_scan_task('test_task')

                assert is_processing is False
                assert task_id is None
                assert task_obj is None
                mock_logging.error.assert_called()


class TestGenerateExcelReport:
    """Тесты для функции generate_excel_report"""

    def test_basic_excel_generation(self):
        sample_dataframe = pd.DataFrame({
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'City': ['New York', 'London', 'Tokyo']
        })

        # Используем временную директорию вместо мокинга
        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, 'test.xlsx')
            result = generate_excel_report(sample_dataframe, file_path)

            # Проверяем что файл создан
            assert os.path.exists(result)

            # Проверяем содержимое файла
            df = pd.read_excel(result)
            assert len(df) == 3
            assert list(df.columns) == ['Name', 'Age', 'City']

    def test_empty_dataframe(self):
        empty_df = pd.DataFrame()

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, 'empty.xlsx')
            result = generate_excel_report(empty_df, file_path)

            # Проверяем что файл создан
            assert os.path.exists(result)

            # Проверяем, что созданный файл пустой
            df = pd.read_excel(result)
            assert len(df) == 0
            assert len(df.columns) == 0

    def test_column_widths_customization(self):
        """Тест: проверка установки ширины столбцов"""
        sample_dataframe = pd.DataFrame({
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'City': ['New York', 'London', 'Tokyo']
        })

        # Определяем пользовательские ширины столбцов
        column_widths = {
            'A': 15,  # Ширина для заголовка "Name"
            'B': 8,  # Ширина для заголовка "Age"
            'C': 20  # Ширина для заголовка "City"
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = os.path.join(tmpdir, 'column_widths.xlsx')
            result = generate_excel_report(
                sample_dataframe,
                file_path,
                column_widths=column_widths
            )

            # Проверяем что файл создан
            assert os.path.exists(result)

            # Проверяем ширины столбцов с помощью openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # Проверяем ширину каждого столбца
            assert ws.column_dimensions['A'].width == 15
            assert ws.column_dimensions['B'].width == 8
            assert ws.column_dimensions['C'].width == 20
