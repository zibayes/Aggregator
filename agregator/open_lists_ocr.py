import json
from datetime import datetime
from typing import List
import redis

import fitz
import pytesseract
from PIL import Image, ImageEnhance
import cv2
import tkinter as tk
from tkinter import filedialog
import numpy as np
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
import os
from pathlib import Path
from transliterate import translit
import Levenshtein
import matplotlib.pyplot as plt
from celery import shared_task
from celery_progress.backend import ProgressRecorder
from .models import OpenLists
from .hash import calculate_file_hash
from .files_saving import save_open_list_files, load_raw_files, delete_files_in_directory

MAX_VAL = 255
UPSCALE = [1]
MONTHS = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06', 'июля': '07',
          'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12', }
pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files/Tesseract-OCR/tesseract.exe'
redis_client = redis.StrictRedis(host='localhost', port=6379, db=0)


def choose_image_file() -> str:
    # file_path = filedialog.askopenfilename(title="Выберите файл изображения", filetypes=[("Изображения", "*.png *.jpg")])
    file_path = filedialog.askdirectory(title="Выберите папку")
    if file_path:
        return file_path


def image_binarization(image_path: str, threshold: int = None) -> cv2.UMat:
    img = cv2.imread(image_path)
    height, width = img.shape[:2]
    if width <= 598 and height <= 845:
        UPSCALE[0] = 8
        img = cv2.resize(img, None, fx=UPSCALE[0], fy=UPSCALE[0], interpolation=cv2.INTER_LANCZOS4)
    img1 = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, bin_img = cv2.threshold(img1, threshold, MAX_VAL, cv2.THRESH_BINARY)
    return img, bin_img


def extract_text_from_image(image: cv2.UMat, psm_conf: str) -> str:
    custom_config = f'--oem 3 --psm {psm_conf}'
    extracted_text = pytesseract.image_to_string(image, lang='rus', config=custom_config)
    return extracted_text


def extract_fio_from_image(image: cv2.UMat, koef: int):
    line_length = 490
    line_color_thresh = 18
    return extract_data_by_lines(image, koef, line_length, line_color_thresh)


def extract_dates_from_image(image: cv2.UMat, koef: int):
    line_length = 112
    line_color_thresh = 18
    return extract_data_by_lines(image, koef, line_length, line_color_thresh)


def extract_data_by_lines(image: cv2.UMat, koef: int, line_length: int, line_color_thresh: int):
    lines = []
    for i in range(0, len(image)):
        j = 0
        while j < len(image[i]) - line_length * koef:
            if image[i][j:j + line_length * koef].sum() / len(image[i][j:j + line_length * koef]) < line_color_thresh:
                # img8[i][j:j+112*4] = np.full(shape=[1,112*4],fill_value=255)
                lines.append((i, j))
                j += line_length * koef
            j += 1
    final_list = []
    for line in lines:
        if not final_list and line[0] - 15 * koef >= 0:
            final_list.append(line)
            continue
        acceptance = []
        for val in final_list:
            if abs(line[0] - val[0]) < koef and abs(line[1] - val[1]) < line_length * koef:
                acceptance.append(False)
            else:
                acceptance.append(True)
        if all(acceptance) and line[0] - 15 * koef >= 0:
            final_list.append(line)
    return final_list


def check_lines(lines, koef: int):
    is_correct = True
    if len(lines) == 6:
        if not 56 * koef <= lines[0][0] <= 71 * koef:
            is_correct = False
        if not 77 * koef <= lines[1][0] <= 93 * koef:
            is_correct = False
        if not 150 * koef <= lines[2][0] <= 225 * koef:
            is_correct = False
        if not 200 * koef <= lines[3][0] <= 300 * koef:
            is_correct = False
        if not 275 * koef <= lines[4][0] <= 450 * koef:
            is_correct = False
        if not 425 * koef <= lines[5][0] <= 780 * koef:  # <= 525 * koef
            is_correct = False
        if not 20 * koef <= lines[1][0] - lines[0][0] <= 24 * koef:
            is_correct = False
        if not 57 * koef <= lines[3][0] - lines[2][0] <= 63 * koef:
            is_correct = False
        if not 152 * koef <= lines[5][0] - lines[4][0] <= 156 * koef:
            is_correct = False
        if is_correct:
            return lines
        return False
    elif len(lines) > 6:
        new_list = []
        for line in lines:
            if len(new_list) == 0 and 56 * koef <= line[0] <= 71 * koef:
                new_list.append(line)
                continue
            if len(new_list) == 1 and 77 * koef <= line[0] <= 93 * koef and 20 * koef <= line[0] - new_list[0][
                0] <= 24 * koef:
                new_list.append(line)
                continue
            if len(new_list) == 2 and 150 * koef <= line[0] <= 225 * koef:
                new_list.append(line)
                continue
            if len(new_list) == 3 and 200 * koef <= line[0] <= 250 * koef and 57 * koef <= line[0] - new_list[2][
                0] <= 63 * koef:
                new_list.append(line)
                continue
            if len(new_list) == 4 and 275 * koef <= line[0] <= 325 * koef:
                new_list.append(line)
                continue
            if len(new_list) == 5 and 425 * koef <= line[0] <= 525 * koef and 152 * koef <= line[0] - new_list[4][
                0] <= 156 * koef:
                new_list.append(line)
                break
        if len(new_list) == 6:
            return new_list
        return False


def cut_dates_from_image(image: cv2.UMat, final_list: List, koef: int):
    line_length = 130
    dates = []
    for date in final_list:
        img = image[date[0] - 15 * koef:date[0], date[1] + 20:date[1] + line_length * koef]
        dates.append(img)
        '''
        plt.imshow(dates[-1], cmap='gray')
        plt.show()
        '''
    return dates


def cut_fio_from_image(image: cv2.UMat, final_list: List, koef: int):
    line_length = 490
    pieces = []
    for date in final_list:
        img = image[date[0] - 18 * koef:date[0], date[1] + 20:date[1] + line_length * koef]
        pieces.append(img)
        '''
        plt.imshow(pieces[-1], cmap='gray')
        plt.show()
        '''
    return pieces


def date_check(date: str) -> bool:
    date_form = re.search(r'«*\d+»* *.[^ ]+ \d{4}', date, re.IGNORECASE)
    if date_form:
        date_form = date_form.group(0)
        month = re.search(r'[а-яА-ЯёЁ]+', date_form)
        if month:
            month = month.group(0)
            if month in MONTHS.keys():
                numbers = re.findall(r'\d+', date_form)
                if len(numbers) == 2 and len(numbers[0]) <= 2 and len(numbers[1]) == 4:
                    if 0 <= int(numbers[0]) <= 31 and 1980 <= int(numbers[1]) <= 2099:
                        return True
    return False


def date_to_dots_format(date: str) -> str:
    day = re.search(r'\b\d{2}\b', date)
    if not day:
        day = re.search(r'\b\d\b', date)
    if day:
        day = day.group(0)
        if len(day) < 2:
            day = '0' + day
    else:
        day = '-'
    year = re.search(r'\d{4}', date)
    if year:
        year = year.group(0)
    else:
        year = '-'
    month = re.search(r'[а-яА-ЯёЁ]+', date)
    if month:
        month = month.group(0)
    if month in MONTHS.keys():
        month = MONTHS[month]
        # date = date.replace(month, '').replace('  ', '.' + months[month] + '.')
    else:
        min_dist = 99
        min_dist_month = ''
        for month_in_list in MONTHS.keys():
            dist = Levenshtein.distance(month, month_in_list)
            if dist < min_dist and dist < 5:
                min_dist = dist
                min_dist_month = MONTHS[month_in_list]
        if min_dist_month:
            month = min_dist_month
    return day + '.' + month + '.' + year


def get_gaps(image: cv2.UMat, koef: int, thresh: int) -> List:
    img = image.copy()
    i = 0
    gaps = []
    while i < len(image):
        if img[i:i + 11 * koef].sum() / (len(img[i]) * 11 * koef) > thresh:
            # img[i:i+11*koef] = np.full(shape=[11*koef,len(img[i])],fill_value=0)
            gaps.append(i)
            i += 11 * koef
        i += 1
    # cv2.imwrite(folder + "/gaps.png", img)
    '''
    plt.imshow(img, cmap='gray')
    plt.show()
    '''
    return gaps


def change_img_perspect(img, dst_pts, src_pts=None):
    if src_pts is None:
        h, w = img.shape[:2]
        src_pts = np.array([[0, 0], [w, 0], [0, h], [w, h]], dtype=np.float32)
    return cv2.warpPerspective(src=img,
                               # матрица преобразования src_pts -> dst_pts
                               M=cv2.getPerspectiveTransform(src_pts, dst_pts),
                               dsize=(w, h))


@shared_task(bind=True)
def process_open_lists(self, uploaded_files, user_id):
    uploaded_files = load_raw_files(uploaded_files, user_id)
    total_processed = [0]
    progress_recorder = ProgressRecorder(self)
    progress_recorder.set_progress(total_processed[0], 0, '')
    open_lists_ids, pages_count, origin_filenames = save_open_list_files(uploaded_files, user_id)
    delete_files_in_directory('uploaded_files/users/' + str(user_id), uploaded_files)
    folder = 'uploaded_files/'
    table = None
    already_uploaded = []
    open_lists = []
    file_groups = {}
    for open_list_id in open_lists_ids:
        open_list = OpenLists.objects.get(id=open_list_id)
        open_lists.append(open_list)
        file = {'path': folder + open_list.source.name, 'origin_name': origin_filenames[str(open_list_id)],
                'processed': 'False', 'pages': {'processed': '0', 'all': pages_count[str(open_list_id)]}}
        file_groups[str(open_list_id)] = file
    progress_json = {'user_id': user_id, 'file_groups': file_groups, 'file_types': 'open_lists',
                     'time_started': datetime.now().strftime(
                         "%Y-%m-%d %H:%M:%S")}
    redis_client.set(self.request.id, json.dumps(progress_json))
    progress_recorder.set_progress(total_processed[0], sum(pages_count.values()), progress_json)
    for open_list in open_lists:
        progress_json['file_groups'][str(open_list.id)]['processed'] = 'Processing'
        if not open_list.source.path.endswith('.pdf'):
            continue
        time_on_start = datetime.now()
        new_table = open_list_ocr(open_list.source.path, progress_recorder, pages_count,
                                  total_processed, open_list.id, progress_json, self.request.id, time_on_start)
        progress_json['file_groups'][str(open_list.id)]['pages']['processed'] = \
            progress_json['file_groups'][str(open_list.id)]['pages']['all']
        progress_json['file_groups'][str(open_list.id)]['processed'] = 'True'
        redis_client.set(self.request.id, json.dumps(progress_json))
        progress_recorder.set_progress(total_processed[0], sum(pages_count.values()), progress_json)
        if isinstance(new_table, pd.DataFrame):
            if table is not None:
                table = table._append(new_table, ignore_index=True)
            else:
                table = new_table
        else:
            already_uploaded.append(file)
    if isinstance(table, pd.DataFrame):
        table = table['Номер листа'].tolist()
    progress_json['time_ended'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return progress_json
    # return table


def open_list_ocr(pdf_path, progress_recorder, pages_count, total_processed,
                  open_list_id, progress_json, task_id, time_on_start):
    open_lists = OpenLists.objects.all()
    for open_list in open_lists:
        if open_list.source and open_list.id != open_list_id and os.path.isfile(open_list.source.path):
            file_hash = calculate_file_hash(pdf_path)
            open_list_hash = calculate_file_hash('uploaded_files/' + open_list.source.name)
            if file_hash == open_list_hash:
                raise FileExistsError(
                    f"Такой файл уже загружен в систему: {progress_json['file_groups'][str(open_list_id)]['origin_name']}")

    false_date = []
    image = None
    image_filename = None
    document = fitz.open(pdf_path)
    for page_number in range(len(document)):
        pages_processed = total_processed[0] + page_number
        progress_json['file_groups'][str(open_list_id)]['pages']['processed'] = page_number
        expected_time = ((datetime.now() - time_on_start) / (pages_processed if pages_processed > 0 else 1)) * (sum(
            pages_count.values()) - pages_processed)
        total_seconds = int(expected_time.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        progress_json['expected_time'] = f"{hours:02}:{minutes:02}:{seconds:02}"
        redis_client.set(task_id, json.dumps(progress_json))
        progress_recorder.set_progress(pages_processed, sum(pages_count.values()),
                                       progress_json)
        # page = document[page_number]
        page = document.load_page(page_number)

        pix = page.get_pixmap(dpi=300)  # Получаем картинку страницы
        image_filename = translit(pdf_path[:pdf_path.rfind(".")], 'ru', reversed=True)
        image_filename = image_filename + ".png"
        pix.save(image_filename)

        '''
        with Image(filename=pdf_path) as img:
            img.save(filename=file[:file.rfind(".")] + f".png")
        '''

        '''
        image_list = page.get_images(full=True)
        for img_index, img in enumerate(image_list):
            img_index = img[0]
            base_image = document.extract_image(img_index)
            image = base_image["image"]
            image_filename = f"page_{page_number + 1}_img_{img_index}.png"
            with open(folder + "/" + image_filename, "wb") as img_file:
                img_file.write(image)
            if image:
                break
        '''
        list_data = {'Номер листа': '', 'Держатель': '', 'Объект': '', 'Работы': '', 'Начало срока': '',
                     'Конец срока': ''}
        # for thresh in range(100, 176, 25):
        binarization_threshold = 120
        imgz, image = image_binarization(image_filename, binarization_threshold)  # , thresh
        height, width = image.shape
        ratio = (width / 596 + height / 842) / 2  # image ratio for different resolutions
        koef = int(UPSCALE[0] * ratio)
        frame = image[204 * koef:778 * koef, 63 * koef:555 * koef]
        frame_rgb = imgz[204 * koef:778 * koef, 63 * koef:555 * koef]
        # cv2.imwrite(folder + "/frame.png", imgz[204*koef:778*koef, 63*koef:555*koef])
        list_number = fio = fio_sklon = object = works = dates = None

        no_lines = []
        line_length = 485
        line_color_thresh = 80
        lines = extract_data_by_lines(frame, koef, line_length, line_color_thresh)
        len_lines = len(lines)
        lines = check_lines(lines, koef)
        while not lines and len_lines <= 10 and binarization_threshold <= 200:
            if binarization_threshold <= 200:
                binarization_threshold += 10
                imgz, image = image_binarization(image_filename, binarization_threshold)
                frame = image[204 * koef:778 * koef, 63 * koef:555 * koef]
            else:
                line_color_thresh += 5
            lines = extract_data_by_lines(frame, koef, line_length, line_color_thresh)
            len_lines = len(lines)
            lines = check_lines(lines, koef)
        if not lines:
            lines = []
        if 5 <= len(lines) <= 6:
            print(lines)
            object = frame_rgb[lines[1][0] + 30 * koef:lines[2][0]]
            # cv2.imwrite(folder + "/object_frame.png", object)
            fio = frame_rgb[lines[2][0] + 30 * koef:lines[3][0]]
            # cv2.imwrite(folder + "/fio_frame.png", fio)
            works = frame_rgb[lines[3][0] + 32 * koef:lines[4][0]]
            # cv2.imwrite(folder + "/works_frame.png", works)
            dates = frame[lines[4][0] + 35 * koef:lines[4][0] + 90 * koef, 200 * koef:]
            dates_rgb = frame_rgb[lines[4][0] + 35 * koef:lines[4][0] + 90 * koef, 200 * koef:]
            # cv2.imwrite(folder + "/dates_frame.png", dates)
        else:
            no_lines.append(image_filename)
            gaps = get_gaps(frame, koef, 250)
            index = 0
            for i in range(len(gaps) - 1):
                if (gaps[i + 1] - gaps[i] < 12 * koef) or (i == 0 and gaps[i] < 12 * koef):
                    continue
                if index == 0 and gaps[i] > 40 * koef:
                    list_number = frame[:gaps[i] + 5 * koef]
                    if list_number.size == 0:
                        list_number = None
                if index == 1 and gaps[i + 1] - gaps[i] > 10 * koef:
                    fio_sklon = frame[gaps[i]:gaps[i + 1]]
                    if fio_sklon.size == 0:
                        fio_sklon = None
                elif index == 3 and gaps[i + 1] - gaps[i] > 22 * koef:
                    object = frame[gaps[i] + 22 * koef:gaps[i + 1] - 4 * koef]
                    if object.size == 0:
                        object = None
                elif index == 5 and gaps[i + 1] - gaps[i] > 14 * koef:
                    fio = frame[gaps[i]:gaps[i + 1] - 5 * koef]
                    if fio.size == 0:
                        fio = None
                elif index == 6 and gaps[i + 1] - gaps[i] > 22 * koef:
                    works = frame[gaps[i] + 23 * koef:gaps[i + 1]]
                    if works.size == 0:
                        works = None
                elif index == 8 and 30 * koef > gaps[i + 1] - gaps[i] > 11 * koef:
                    dates = frame[gaps[i] + 8 * koef:gaps[i + 1] + 4 * koef]
                    dates_rgb = frame_rgb[gaps[i] + 8 * koef:gaps[i + 1] + 4 * koef]
                    if dates.size == 0:
                        dates = None
                index += 1
                '''
            if list_number is not None and fio is not None and fio_sklon is not None \
                    and object is not None and works is not None and dates is not None:
                cv2.imwrite(folder + "/list_number_gap.png", list_number)
                cv2.imwrite(folder + "/fio_gap.png", fio)
                cv2.imwrite(folder + "/fio_sklon_gap.png", fio_sklon)
                cv2.imwrite(folder + "/object_gap.png", object)
                cv2.imwrite(folder + "/works_gap.png", works)
                cv2.imwrite(folder + "/dates_gap.png", dates)
                '''
        if not (fio is not None and object is not None
                and works is not None and dates is not None):
            fio = image[390 * koef:490 * koef, 60 * koef:560 * koef]
            # cv2.imwrite(folder + "/fio.png", fio)
            '''
            final_list = extract_fio_from_image(fio, koef)
            fios = cut_fio_from_image(fio, final_list, koef)
            if len(fios) > 0:
                fio = fios[0]
            '''

            works = image[415 * koef:580 * koef, 60 * koef:560 * koef]
            gaps = get_gaps(works, koef, 240)
            gap_num = 0
            for i in range(len(gaps) - 1):
                if (gaps[i + 1] - gaps[i] < 36 * koef) or (i == 0 and gaps[i] < 36 * koef):
                    continue
                gap_num = i
                break
            works = imgz[415 * koef:580 * koef, 60 * koef:560 * koef]
            if len(gaps) > 1:
                works = works[gaps[gap_num] + 24 * koef:gaps[gap_num + 1] if len(gaps) > 1 else len(works)]
            if works.size > 0:
                pass
                # cv2.imwrite(folder + "/works.png", works)
            else:
                works = imgz[415 * koef:580 * koef, 60 * koef:560 * koef]

            object = image[295 * koef:450 * koef, 60 * koef:560 * koef]
            gaps = get_gaps(object, koef, 250)
            gap_num = 0
            for i in range(len(gaps) - 1):
                if (gaps[i + 1] - gaps[i] < 21 * koef) or (i == 0 and gaps[i] < 21 * koef):
                    continue
                gap_num = i
                break
            object = imgz[295 * koef:450 * koef, 60 * koef:560 * koef]
            if len(gaps) > 1:
                object = object[
                         gaps[gap_num] + 24 * koef:(gaps[gap_num + 1] if len(gaps) > 1 else len(object)) - 5 * koef]
            if object.size > 0:
                pass
                # cv2.imwrite(folder + "/object.png", object)
            else:
                object = imgz[295 * koef:450 * koef, 60 * koef:560 * koef]

            dates = image[515 * koef:600 * koef, 260 * koef:540 * koef]
            dates_rgb = imgz[515 * koef:600 * koef, 260 * koef:540 * koef]
        if not list_data['Номер листа']:
            list_number = imgz[196 * koef:235 * koef, 200 * koef:420 * koef]
            # cv2.imwrite(folder + "/list_number.png", list_number)
            extracted_text = extract_text_from_image(list_number, '1').upper().replace('О', '0') \
                .replace('()', '0').replace('З', '3')
            list_number = re.search(r'№[ \n]*.*\d+-\d+.*', extracted_text, re.IGNORECASE)
            if list_number:
                list_data['Номер листа'] = list_number.group(0)
            else:
                list_data['Номер листа'] = extracted_text
        progress_recorder.set_progress(pages_processed + 1 / 5, sum(pages_count.values()),
                                       progress_json)
        if not list_data['Держатель']:
            extracted_text = extract_text_from_image(fio, '1')
            list_holder = re.search(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+', extracted_text)
            if not list_holder or len(list_holder.group(0)) <= 12:
                fio_sklon = imgz[245 * koef:285 * koef, 182 * koef:437 * koef]
                # cv2.imwrite(folder + "/fio_sklon.png", fio_sklon)
                extracted_text = extract_text_from_image(fio_sklon, '1')
                list_holder = re.search(r'[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+[ \n]+[А-ЯЁ]+[а-яё]+',
                                        extracted_text)
            if list_holder:
                list_data['Держатель'] = list_holder.group(0)
            else:
                list_data['Держатель'] = extracted_text
        progress_recorder.set_progress(pages_processed + 2 / 5, sum(pages_count.values()),
                                       progress_json)
        if not list_data['Объект']:
            extracted_text = extract_text_from_image(object, '1')
            list_data['Объект'] = extracted_text
        progress_recorder.set_progress(pages_processed + 3 / 5, sum(pages_count.values()),
                                       progress_json)
        if not list_data['Работы']:
            extracted_text = extract_text_from_image(works, '1')
            list_data['Работы'] = extracted_text
        progress_recorder.set_progress(pages_processed + 4 / 5, sum(pages_count.values()),
                                       progress_json)
        if not list_data['Начало срока'] or not list_data['Конец срока']:
            dates_list = extract_dates_from_image(dates, koef)
            # imgz_dates = dates # imgz[515*koef:600*koef, 260*koef:540*koef]
            dates_list = cut_dates_from_image(dates_rgb, dates_list, koef)
            period_dates = []
            date_index = 0
            extracted_text = None
            all_sucess = [False for i in dates_list]
            if len(dates_list) > 2:  # if dates_list == 3
                i = 0
                for date in dates_list:
                    if i > 1 and all(all_sucess[:2]):
                        break
                    extracted_text = extract_text_from_image(date,
                                                             '6')  # .replace('20242.', '2024  г.').replace('/', '1').replace('З', '3').replace('[', '')
                    date_form = re.search(r'«*\d+»* *.[^ ]+ \d{4}', extracted_text, re.IGNORECASE)
                    if date_form:
                        date_form = date_form.group(0)
                    if date_form and date_check(date_form):
                        print('first', i)
                        period_dates.append(date_form)
                        # cv2.imwrite(folder + f"/date{i}.png", date)
                        all_sucess[i] = True
                    else:
                        j = 0
                        date_is_correct = False
                        while j < 6 and (not date_form or not date_is_correct):
                            if j == 1:
                                h, w = date.shape[:2]
                                date = change_img_perspect(date, dst_pts=np.array(
                                    [[-0.02 * w, 0], [0.99 * w, 0], [0, h], [w, h]], dtype=np.float32))
                            if j > 1 or j == 0:
                                date = cv2.convertScaleAbs(date, alpha=1.2, beta=0)
                            extracted_text = extract_text_from_image(date, '6').replace('20242.', '2024  г.').replace(
                                '/', '1').replace('З', '3').replace('2924', '2024')
                            date_form = re.search(r'«*\d+»* *.[^ ]+ \d{4}', extracted_text, re.IGNORECASE)
                            if date_form:
                                date_form = date_form.group(0)
                                date_is_correct = date_check(date_form)
                            j += 1
                        print('sec', i, j)
                        period_dates.append(extracted_text)
                        # cv2.imwrite(folder + f"/date{i}.png", date)
                        if date_is_correct:
                            all_sucess[i] = True
                    i += 1
                print('cut =', period_dates)
            else:
                extracted_text = extract_text_from_image(dates, '6')
                '''
                substr_place = extracted_text.find('Срок')
                if substr_place > 0:
                    initial_str = 'Срок действия открытого листа'
                    distance = Levenshtein.distance(extracted_text[substr_place:substr_place+len(initial_str)], initial_str)
                    # [А-Яа-яёЁA-Za-z \n,0-9:.;"()«»\\–-]+?(?=Дата)
                    if distance < 12:
                '''
                extracted_text = extracted_text.replace('20242.', '2024  г.').replace('/', '1').replace('З',
                                                                                                        '3').replace(
                    '[', '')
                period_dates = re.findall(r'«*\d+»* *.[^ ]+ \d{4}', extracted_text, re.IGNORECASE)
                print('re =', period_dates)
            if len(period_dates) > 0:
                date = period_dates[0]
                if not re.search(r'«*\d+»* *.[^ ]+ \d{4}', date, re.IGNORECASE) and len(period_dates) > 2:
                    date = period_dates[2]
                if date:
                    list_data['Начало срока'] = date_to_dots_format(date)
            if len(period_dates) > 1:
                date = period_dates[1]
                if date:
                    list_data['Конец срока'] = date_to_dots_format(date)
            '''
            if all(list_data.values()):
                break
            '''
        total_processed[0] += len(document)
        os.remove(image_filename)
        open_list = OpenLists.objects.get(id=open_list_id)
        open_list.number = list_data['Номер листа']
        open_list.holder = list_data['Держатель']
        open_list.object = list_data['Объект']
        open_list.works = list_data['Работы']
        open_list.start_date = list_data['Начало срока']
        open_list.end_date = list_data['Конец срока']
        open_list.save()

        table_path = "uploaded_files/open_lists/Открытые листы.xlsx"
        df_new = pd.DataFrame(list_data, columns=list_data.keys(), index=[0])
        table_data = df_new
        if os.path.exists(table_path):
            df_existing = pd.read_excel(table_path)
            df_new = df_existing._append(df_new, ignore_index=True)
        '''
        df_new['Начало срока'] = pd.to_datetime(df_new['Начало срока'], format='%d.%m.%Y', dayfirst=True)
        df_new.sort_values(by='Начало срока', ascending=False, inplace=True)
        df_new['Начало срока'] = df_new['Начало срока'].dt.strftime('%d.%m.%Y')
        '''
        with pd.ExcelWriter(table_path) as writer:
            df_new.to_excel(writer, sheet_name="Sheet1", index=False)
        wb = load_workbook(table_path)
        ws = wb.active
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 100
        ws.column_dimensions['D'].width = 100
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        font = Font(
            name='Times New Roman',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
        )
        {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}
        for i in range(1, len(df_new.values) + 2):
            if i == 1:
                ws.row_dimensions[0].height = 50
            else:
                ws.row_dimensions[i].height = 80
            for cell in ws[i]:
                if cell.value:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wb.save(table_path)
        return table_data


@shared_task
def error_handler_open_lists(task, exception, exception_desc):
    print(f"Задача {task.id} завершилась с ошибкой: {exception} {exception_desc}")
    progress_json = json.loads(redis_client.get(task.id))
    for open_list_id, source in progress_json['file_groups'].items():
        print(open_list_id, source)
        if source['processed'] != 'True':
            open_list = OpenLists.objects.get(id=open_list_id)
            open_list.delete()
    progress_json['time_ended'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    raise type(exception)({"error_text": str(exception), "progress_json": progress_json}) from exception


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    file_path = choose_image_file()
    dir = file_path
    for root, dirs, files in os.walk(dir):
        for file in files:
            if not file.lower().endswith('.pdf'):
                continue
            print(file)
            pdf_path = os.path.join(f, file)
            open_list_ocr(file_path, pdf_path)
